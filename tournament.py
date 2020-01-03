#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import numpy as np
from openpyxl import utils
import copy
import re
import slackweb
import time

#トーナメント処理のクラス
class Tournament():
    def __init__(self):
        #罫線の位置と色の指定→(色_引く位置)
        self.black='000000'
        self.black_right = Border(right=Side(style='thin', color=self.black))
        self.black_left = Border(left=Side(style='thin', color=self.black))
        self.black_top_right = Border(top=Side(style='thin', color=self.black),
                        right=Side(style='thin', color=self.black)
        )
        self.black_under_right = Border(bottom=Side(style='thin', color=self.black),
                        right=Side(style='thin', color=self.black)
        )
        self.black_top_left = Border(top=Side(style='thin', color=self.black),
                        left=Side(style='thin', color=self.black)
        )
        self.black_under_left = Border(bottom=Side(style='thin', color=self.black),
                        left=Side(style='thin', color=self.black)
        )
        self.black_under=Border(bottom=Side(style='thin', color=self.black))
        self.black_under_left_right = Border(bottom=Side(style='thin', color=self.black),
                        left=Side(style='thin', color=self.black),
                        right=Side(style='thin', color=self.black)
        )
        self.black_top_left_right = Border(top=Side(style='thin', color=self.black),
                        left=Side(style='thin', color=self.black),
                        right=Side(style='thin', color=self.black)
        )
        self.black_left_right = Border(left=Side(style='thin', color=self.black),
                        right=Side(style='thin', color=self.black)
        )
        self.right_center=Alignment(horizontal = 'right',
                                vertical = 'center')
        self.left_center=Alignment(horizontal = 'left',
                                vertical = 'center')
        self.center_bottom=Alignment(horizontal = 'center',
                                vertical = 'bottom')
    
    #Excelファイルの読み込み
    def file_load(self,load_file):
        ans=openpyxl.load_workbook(load_file)
        return ans
    
    #列データからデータをリスト化
    def columuns_list_making(self,num):
        ans_list=[]
        for ans_list_obj in list(self.sheet.columns)[num]:
                ans_list.append(ans_list_obj.value)
        ans_list.pop(0)
        return ans_list
    
    #日付データの文字列化(Excelでの表記からの変更)
    def date_to_string(self,data):
        for i in range(len(data)):
            data[i] = utils.datetime.from_excel(data[i])
            data[i]=data[i].strftime('%m%d')
        return data
    
    #リストを種類ごとに並べてソートする
    def list_to_category(self,data):
        copy = copy.deepcopy(data)
        category=list(set(copy))
        category.sort()
        return category

    #リストの要素の足し算
    def list_add(self,a,b):
        return [x + y for (x, y) in zip(a,b)]

    #リスト同士のすべての組み合わせを作る
    def all_combine(self,a,b):
        ans_list=[]
        for i in range(len(a)):
            for j in range(len(b)):
                ans_list.append(b[j]+a[i])
        return ans_list

    #必要数のシート作成
    def create_need_sheet(self,data):
        for i in range(len(data)):
            self.data_sheet.create_sheet(index=i+1,title=data[i])

    #メンバーリストの作成
    def making_memberlist(self):
        for i in range(len(self.rank_date_category)):
            sheet=self.data_sheet[self.rank_date_category[i]]
            sheet["A1"]="ランク"
            sheet["B1"]="日付"
            sheet["C1"]="場所"
            sheet["D1"]="種目"
            sheet["E1"]="名前"
            sheet["F1"]="サークル名"
            sheet["G1"]="学年"
            sheet["H1"]="強さ"
            #個人のデータをトーナメントごとに入れていく
            for j in range(len(self.rank_date_list)):
                max_row=sheet.max_row
                if self.rank_date_list[j]==self.rank_date_category[i]:
                    sheet.cell(row=max_row+1,column=5).value = self.name[j]
                    sheet.cell(row=max_row+1,column=6).value = self.circle[j]
                    sheet.cell(row=max_row+1,column=7).value = self.grade[j]
                    sheet.cell(row=max_row+1,column=8).value = 0
                    sheet["A2"]=self.rank[j]
                    sheet["B2"]=self.date[j]
            #参加者がいないトーナメントは削除する
            if sheet.max_row==1:
                    self.data_sheet.remove(self.data_sheet[self.rank_date_category[i]])
            sheet["D2"]=self.category

    #memberlistの作成
    def main_memberlist(self,load_file,category,place):
        self.category=category
        self.place=place
        #datasheetから、ランク、日付、名前、サークル名、学年を取得
        self.data_sheet=self.file_load(load_file)
        self.sheet=self.data_sheet['Sheet1']
        self.rank=self.columuns_list_making(0)
        self.date=self.columuns_list_making(1)
        self.name=self.columuns_list_making(2)
        self.circle=self.columuns_list_making(3)
        self.grade=self.columuns_list_making(4)
        #日付データを文字列に変更
        self.date=self.date_list_to_string(self.date)
        #被りのある日付データ、ランクデータを種類ごとのデータに変更
        self.date_category=self.list_to_category(self.date)
        self.rank_category=self.list_to_category(self.rank)
        #日付とランクの組み合わせデータを作成
        self.rank_date_list=self.list_add(self.rank,self.date)
        #日付とランクの組み合わせの種類データを作成(トーナメント数に対応)
        self.rank_date_category=self.all_combine(self.date_category,self.rank_category)
        #必要なトーナメント数分のシートを作成
        self.create_need_sheet(self.rank_date_category)
        self.making_memberlist()
        self.data_sheet.remove(self.data_sheet['Sheet1'])
        self.data_sheet.save('memberlist.xlsx')

    #左右の2番目のブロックの作図
    def second_block(self,n):
        for i in range(2**(n-1)):
            self.tur['C'+str(9+i*16)].border=self.black_top_right
            self.tur['C'+str(16+i*16)].border=self.black_under_right
            self.tur[chr(ord('C')+n*2+1)+str(9+i*16)].border=self.black_top_left
            self.tur[chr(ord('C')+n*2+1)+str(16+i*16)].border=self.black_under_left
            for j in range(6):
                self.tur['C'+str(10+i*16+j)].border=self.black_right
                self.tur[chr(ord('C')+n*2+1)+str(10+i*16+j)].border=self.black_left
    
    #最後を除く中間のブロックの作図
    def except_block(self,n):
        cnt=2
        while n-cnt>=0:
            for i in range(2**(n-cnt)):
                self.tur[chr(ord('C')+cnt-1)+str((2**(cnt+1))+5+i*(2**(cnt+3)))].border=self.black_top_left_right
                self.tur[chr(ord('C')+cnt-1)+str((2**(cnt+1))+5+(2**(cnt+2))-1+i*(2**(cnt+3)))].border=self.black_under_left_right
                self.tur[chr(ord('C')+2*n-cnt+2)+str((2**(cnt+1))+5+i*(2**(cnt+3)))].border=self.black_top_left_right
                self.tur[chr(ord('C')+2*n-cnt+2)+str((2**(cnt+1))+5+(2**(cnt+2))-1+i*(2**(cnt+3)))].border=self.black_under_left
                for j in range(2**(cnt+2)-2):
                    self.tur[chr(ord('C')+cnt-1)+str((2**(cnt+1))+6+i*(2**(cnt+3))+j)].border=self.black_right
                    self.tur[chr(ord('C')+2*n-cnt+2)+str((2**(cnt+1))+6+i*(2**(cnt+3))+j)].border=self.black_left
            cnt+=1

    #最後のブロックの作図
    def final_block(self,n):
        num=2**(n-1)
        row=num*8+4
        self.tur[chr(ord('C')+n)+str(row)].border=self.black_under_left_right
        self.tur[chr(ord('C')+n+1)+str(row)].border=self.black_under_left_right
        self.tur[chr(ord('C')+n)+str(row-1)].border=self.black_left_right
        self.tur[chr(ord('C')+n+1)+str(row-1)].border=self.black_left_right
    
    #左右の端以外の作図(デフォルト)
    def default(self,n):
        self.second_block(n)
        self.except_block(n)
        self.final_block(n)
    
    #ランク、日付、場所、シングルスorダブルス、名前、サークル名、学年、強さをリスト化
    #表記名をリスト化　表記上段→名前　表記下段→(サークル名+学年)
    def make_list(self):
        self.write_name=[]
        self.write_belong=[]
        self.rank=self.sheet["A2"].value
        self.date=self.sheet["B2"].value
        self.place=self.sheet["C2"].value
        self.category=self.sheet["D2"].value
        self.name=self.columuns_list_making(4)
        self.circle=self.columuns_list_making(5)
        self.grade=self.columuns_list_making(6)
        self.grade = [5 if i == "M1" else i for i in self.grade]
        self.grade = [6 if i == "M2" else i for i in self.grade]
        self.grade = [7 if i == "D1" else i for i in self.grade]
        self.grade = [8 if i == "D2" else i for i in self.grade]
        self.strength=self.columuns_list_making(7)
        if self.category=='シングルス':
            for i in range(len(self.name)):
                self.write_name.append(self.name[i])
                self.write_belong.append("("+self.circle[i]+str(self.grade[i])+")")

        #学年と強さからポイントをつけて行く
        self.point= [x*5 + y for (x, y) in zip(self.grade, self.strength)]

    def sort_list(self,data):
        ans=[]
        for i in range(len(self.sortindex)):
            ans.append(data[self.sortindex[i]])
        ans.reverse()
        for i in range(len(self.sortindex)):
            data[i]=ans[i]


    def decide_right_index(self):
        if self.cnt==7:
            self.right_index='P'
            self.right_index2='O'

        if self.cnt==6:
            self.right_index='N'
            self.right_index2='M'
        if self.cnt==5:
            self.right_index='L'
            self.right_index2='K'
        if self.cnt==4:
            self.right_index='J'
            self.right_index2='I'
        if self.cnt==3:
            self.right_index='H'
            self.right_index2='G'

    #シードの位置を決める
    def seed_place(self):

        if self.seed_num!=0:
            #一つ小さいトーナメント表のどこにシードが入るかを記録するリストを用意する
            #例えば、53人のトーナメント表だと32人のトーナメント表を考える
            before_seed_place=[0]*self.defo_num
            div=int(self.defo_num/self.seed_num)
            std=min(self.seed_num-self.defo_num%self.seed_num,self.defo_num%self.seed_num)
            #シードと次のシードの間に何人がいるかを記録するリストの用意
            if std==self.defo_num%self.seed_num:
                seed_space_num=[div]*self.seed_num
            else:
                seed_space_num=[div+1]*self.seed_num

            for i in range(std):
                if std==self.defo_num%self.seed_num:
                    seed_space_num[int(self.seed_num/std)*(i+1)-1]=div+1
                else:
                    seed_space_num[int(self.seed_num/std)*(i+1)-1]=div
            counter=-1
            for i in range(self.seed_num):
                counter+=seed_space_num[i]
                before_seed_place[counter]=1
            counter=0
            for i in range(self.defo_num):
                if before_seed_place[i]==0:
                    counter+=2
                else:
                    counter+=1
                    self.seed.append(counter)
                    self.number[counter-1]=1

    #トーナメント票の左右の人数を決定する
    def left_right_num(self):
        cntseed=0
        for i in range(1,int((self.num+1)/2)+1):
            for j in range(self.seed_num):
                if self.seed[j]==i:
                    cntseed+=1

        if (int((self.num+1)/2)-cntseed)%2==0:
            self.left_number=int((self.num+1)/2)
            if self.left_number==cntseed:
                self.left_number-=1
            self.right_number=self.num-self.left_number

        else:
            self.left_number=int((self.num+1)/2)-1
            self.right_number=self.num-self.left_number

    #シードを考慮したブロックを作る
    def make_block(self):
        counting=0
        for i in range(len(self.circle)):
            if self.number[i]==1:
                counting+=2
            else:
                counting+=1
            if counting==4:
                counting=0
                self.block.append(i)

    #シード選手の情報をまとめる
    def seed_info_list_make(self,data):
        ans=[]
        for i in range(len(self.seed)):
             ans.append(data[i])
        for i in range(len(self.seed)):
             data.pop(0)
        return ans

    #シード選手の情報をまとめると同時にシード以外の選手をまとめる
    def seed_prepare(self):
        self.seed_name=self.seed_info_list_make(self.name)
        self.seed_circle=self.seed_info_list_make(self.circle)
        self.seed_grade=self.seed_info_list_make(self.grade)
        self.seed_write_name=self.seed_info_list_make(self.write_name)
        self.seed_write_belong=self.seed_info_list_make(self.write_belong)
    
    #バップの被りを考慮して被りが多い順にトーナメント表に入れるように並べ替える
    #最初にシードを入れることも考慮して最終の順番を決定する
    def decide_order(self):
        bup_list=[]
        for i in range(self.bup_data.max_column):
            for bup_obj in list(self.bup_data.columns)[i]:
                if bup_obj.value!=None:
                    bup_list.append(bup_obj.value)

        #幾つのバップに入っているかをカウント
        self.count_bup=[0]*len(self.circle)
        for i in range(len(self.circle)):
            for j in range(len(bup_list)):
                if self.circle[i]==bup_list[j]:
                    self.count_bup[i]+=1

        #ポイント順にソートした後のindexを返す
        self.sortindex = np.argsort(self.count_bup)
        self.sort_list(self.name)
        self.sort_list(self.circle)
        self.sort_list(self.grade)
        self.sort_list(self.write_name)
        self.sort_list(self.write_belong)
        #最終的な順番の決定
        self.name=self.seed_name+self.name
        self.circle=self.seed_circle+self.circle
        self.grade=self.seed_grade+self.grade
        self.write_belong=self.seed_write_belong+self.write_belong
        self.write_name=self.seed_write_name+self.write_name


    #重みの計算、バップ被りのサークルが近いところにあると重みが加算される
    def weight_caluculation(self,i,y):
        for j in range(len(self.circle)):
            self.weight[j][y]+=10000
        list_bap=[0]*self.bup_data.max_column
        weight_circle=[]
        #バップのある組み合わせを決定する
        for j in range(self.bup_data.max_column):
            for bap_obj in list(self.bup_data.columns)[j]:
                if bap_obj.value!=None:
                    if self.circle[i]==bap_obj.value:
                        list_bap[j]=1
        #バップのある組み合わせから全サークルをリスト化
        for j in range(self.bup_data.max_column):
            if list_bap[j]==1:
                for bap_obj in list(self.bup_data.columns)[j]:
                    if bap_obj.value!=None:
                        weight_circle.append(bap_obj.value)
        #バップのあるサークルを距離に応じた重みをつける
        for j in range(len(self.circle)):
            for k in range(len(weight_circle)):
                if weight_circle[k]==self.circle[j] and i!=j:
                    for l in range(len(self.circle)):
                        self.weight[j][l]+=self.distance(y,l)
        #同名サークルに重みをつける
        for j in range(len(self.circle)):
            if self.circle[i]==self.circle[j] and j!=i:
                for k in range(len(self.circle)):
                    self.weight[j][k]+=5*self.distance(y,k)

    #ブロックの位置を決定する
    def blockcount(self,x):
        ans=0
        for i in range(len(self.block)):
            if x<=self.block[m]:
                ans=i+1
                break
        return ans


    #ブロック同士の距離を決定する
    def distance(self,x,y):
        if int(self.blockcount(x)/8)!=int(self.blockcount(y)/8):
            return 1
        elif int(self.blockcount(x)/4)!=int(self.blockcount(y)/4):
            return 3
        elif int(self.blockcount(x)/2)!=int(self.blockcount(y)/2):
            return 6
        else:
            return 10

    
    def decide_num_list(self,data):
        ans=[]
        for i in range(len(self.number_decide)):
            ans.append(data[self.number_decide[i]])
        for i in range(len(self.number_decide)):
            data[i]=ans[i]

    def bup_operation(self):
        self.decide_order()
        #それぞれの選手をトーナメント表のどこに入れるかを決定するリストの用意
        self.number_decide=[]
        for i in range(len(self.seed)):
            ans=1000000
            for j in range(len(self.circle)):
                for k in range(len(self.seed)):
                    if j==self.seed[k]-1:
                        if ans>self.weight[i][j]:
                            ans=self.weight[i][j]
                            y=j
            self.number_decide.append(y)
            self.weight_caluculation(i,y)

        for i in range(len(self.seed),len(self.circle)):
            ans=1000000
            for j in range(len(self.circle)):
                if ans>self.weight[i][j]:
                    ans=self.weight[i][j]
                    y=j
            self.number_decide.append(y)
            self.weight_caluculation(i,y)
        
        self.decide_num_list(self.name)
        self.decide_num_list(self.circle)
        self.decide_num_list(self.grade)
        self.decide_num_list(self.write_belong)
        self.decide_num_list(self.write_name)

    #トーナメント表にシードを決定後の線を引く
    def final_line_draw(self):
        count_left=8
        count_right=8
        noseed=-1

        if self.cnt>=3:
            for i in range(self.num):
                if self.left_number>=i+1:
                    if self.number[i]==1:
                        self.tur['A'+str(count_left)].value=self.write_name[i]
                        self.tur['A'+str(count_left+1)].value=self.write_belong[i]
                        self.tur['B'+str(count_left)].border=self.black_under
                        count_left+=8
                    elif noseed==-1 and self.number[i]==0:
                        self.tur['A'+str(count_left-2)].value=self.write_name[i]
                        self.tur['A'+str(count_left-1)].value=self.write_belong[i]
                        self.tur['B'+str(count_left-1)].border=self.black_top_right
                        self.tur['B'+str(count_left)].border=self.black_right
                        self.tur.merge_cells('B'+str(count_left)+':'+'B'+str(count_left+1))
                        self.tur['B'+str(count_left)].value=self.game_number
                        self.tur['B'+str(count_left)].border=self.black_right
                        self.game_number+=1
                        noseed*=-1
                    else:
                        self.tur['A'+str(count_left+2)].value=self.write_name[i]
                        self.tur['A'+str(count_left+3)].value=self.write_belong[i]
                        self.tur['B'+str(count_left+2)].border=self.black_under_right
                        self.tur['B'+str(count_left+1)].border=self.black_right
                        count_left+=8
                        noseed*=-1
                else:
                    if self.number[i]==1:
                        self.tur[self.right_index+str(count_right)].value=self.write_name[i]
                        self.tur[self.right_index+str(count_right+1)].value=self.write_belong[i]
                        self.tur[self.right_index2+str(count_right)].border=self.black_under
                        count_right+=8
                    elif noseed==-1 and self.number[i]==0:
                        self.tur[self.right_index+str(count_right-2)].value=self.write_name[i]
                        self.tur[self.right_index+str(count_right-1)].value=self.write_belong[i]
                        self.tur[self.right_index2+str(count_right-1)].border=self.black_top_left
                        self.tur[self.right_index2+str(count_right)].border=self.black_left
                        self.tur.merge_cells(self.right_index2+str(count_right)+':'+self.right_index2+str(count_right+1))
                        self.tur[self.right_index2+str(count_right)].value=self.game_number
                        self.tur[self.right_index2+str(count_right)].border=self.black_left
                        self.game_number+=1
                        noseed*=-1
                    else:
                        self.tur[self.right_index+str(count_right+2)].value=self.write_name[i]
                        self.tur[self.right_index+str(count_right+3)].value=self.write_belong[i]
                        self.tur[self.right_index2+str(count_right+2)].border=self.black_under_left
                        self.tur[self.right_index2+str(count_right+1)].border=self.black_left
                        count_right+=8
                        noseed*=-1
    
    #番号の挿入、縦の列と連結、右寄せ/左寄せ
    def insert_num(self,index,num,game,direction):
        self.tur[index+str(num)].value=game
        self.tur.merge_cells(index+str(num)+':'+index+str(num+1))
        if direction=="right":
            self.tur[index+str(num)].alignment=self.left_center
        if direction=="left":
            self.tur[index+str(num)].alignment=self.right_center

    #番号の挿入、横のセルと連結、中央寄せ
    def insert_num_final(self,index,num,game):
        self.tur[index+str(num)].value=game
        self.tur.merge_cells(index+str(num)+':'+chr(ord(index)+1)+str(num))
        self.tur[index+str(num)].alignment=self.center_bottom

    #試合番号やランクなどの挿入
    def game_number_insert(self):
        left='C'
        right=chr(ord('C')+self.cnt*2-3)
        cnt=self.cnt-3
        count=1
        while cnt>=0:
            for i in range(2**cnt):
                self.insert_num(left,4+2**(count+2)+i*2**(count+3),self.game_number,"left")
                self.game_number+=1
            left=chr(ord(left)+1)
            for i in range(2**cnt):
                self.insert_num(right,4+2**(count+2)+i*2**(count+3),self.game_number,"right")
                self.game_number+=1
            right=chr(ord(right)-1)
            count+=1
            cnt-=1
        self.insert_num_final(left,5+2**(count+1)+i*2**(count+2),self.game_number)
        self.insert_num_final(left,1,self.rank)
        self.insert_num_final(left,3,self.date)
        self.insert_num_final(left,4,self.place)

    #トーナメントの作成
    def make(self):
        self.tournament = openpyxl.Workbook()
        self.member = openpyxl.load_workbook('memberlist.xlsx')
        self.bup = openpyxl.load_workbook("bup.xlsx")
        self.bup_data=self.bup['バップリスト']


        for r in range(len(self.member.sheetnames)):
        #参加者リストの読み込み
        #基本情報入力 ランク　場所　種目
            self.sheet = self.member.worksheets[r]
            self.make_list()
            #ポイント順にソートした後のindexを返す
            self.sortindex = np.argsort(self.point)
            self.sort_list(self.name)
            self.sort_list(self.circle)
            self.sort_list(self.grade)
            self.sort_list(self.write_name)
            self.sort_list(self.write_belong)
            #numに人数を記録
            self.num=self.sheet.max_row-1
            #トーナメント用のシートを用意
            self.tur =self.tournament.create_sheet()
            for i in range(10):
                if self.num<=2**(i+2):
                    self.n=i
                    break
            self.default(self.n)
            self.defo_num=2**(self.n+1)
            self.cnt=self.n+2
            self.decide_right_index()

            self.seed_num=2**self.cnt-self.num

            self.seed=[]
            self.number=[0]*(self.num+10)
            self.seed_place()
            self.left_right_num()

            self.weight=[[0 for i in range(len(self.circle))] for j in range(len(self.circle))]
            self.block=[]
            self.seed_prepare()
            self.bup_operation()
            if r==0:
                predate=self.date
            if self.date!=predate or r==0:
                self.game_number=100
            else:
                self.game_number=100+int(self.game_number/100)*100
            predate=self.date
            self.final_line_draw()
            self.game_number_insert()

            for i in range(500):
                self.tur['A'+str(i+1)].alignment = self.center_bottom
                self.tur[self.right_index+str(i+1)].alignment = self.center_bottom
            
            self.tur.title=self.rank+self.date
        self.tournament.remove(self.tournament.worksheets[0])
        self.tournament.save("完成版.xlsx")

    #トーナメント表から日付をリスト化する
    def date_print(self,tournament_file,member_file):
        self.tournament=openpyxl.load_workbook(tournament_file)
        self.member=openpyxl.load_workbook(member_file)
        date_list=[]
        for r in range(len(tournament.sheetnames)):
            date_list.append(tournament.sheetnames[r][2:6])

        date_list=list(set(date_list))
        date_list.sort()
        return date_list

    #日付ごとトーナメント表/メンバー表の作成
    def date_making(self,tournament_file,member_file,date):
        self.tournament=openpyxl.load_workbook(tournament_file)
        self.member=openpyxl.load_workbook(member_file)

        cnt=0
        for r in range(len(self.tournament.sheetnames)):
            if self.tournament.sheetnames[r+cnt][2:6]!=date:
                self.tournament.remove(self.tournament.worksheets[r+cnt])
                cnt-=1
            if self.member.sheetnames[r+cnt][2:6]!=date:
                self.member.remove(self.member.worksheets[r+cnt])
        self.tournament.save(date+'トーナメント.xlsx')
        self.member.save(date+'メンバー.xlsx')

    #必要分のスコア表の作成
    def list_make(self,index):
        for i in range(6,self.tur.max_row+1):
            if self.tur[index+str(i)].value:
                if int(self.tur[index+str(i)].value)>=100:
                    row=self.opr.max_row+1
                    self.opr["B"+str(row)].value=index
                    self.opr["C"+str(row)].value=i
                    self.opr["D"+str(row)].value=self.tur[index+str(i)].value
                    self.new =openpyxl.load_workbook("バドミントンスコアシート.xlsx")
                    self.score = self.new['シングルス用スコア']
                    self.new.remove(self.new['ダブルス用スコア'])
                    self.score.title=str(self.opr["D"+str(row)].value)
                    self.new.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(self.opr["D"+str(row)].value)+".xlsx")
    
    #試合番号を検索する
    #検索して記録する
    def game_number_search(self):
        left='B'
        right=chr(ord('B')+self.cnt*2-1)
        while left!=chr(ord(right)-1):
            self.list_make(left)
            self.list_make(right)
            left=chr(ord(left)+1)
            right=chr(ord(right)-1)
        self.list_make(left)

    #試合番号表にコピーする
    def game_copy(self,name):
        for i in range(1,self.opr.max_row+1):
            row=self.game.max_row+1
            if self.opr["B"+str(i)].value==name:
                self.game["A"+str(row)]=self.rank
                self.game["B"+str(row)]=name
                self.game["C"+str(row)]=self.opr["C"+str(i)].value
                self.game["D"+str(row)]=self.opr["D"+str(i)].value
                if self.opr["E"+str(i)].value:
                    self.game["E"+str(row)]=self.opr["E"+str(i)].value
                    self.game["F"+str(row)]=self.opr["F"+str(i)].value
                if self.opr["I"+str(i)].value:
                    self.game["G"+str(row)]=self.opr["I"+str(i)].value
                    self.game["H"+str(row)]=self.opr["J"+str(i)].value

    #試合の設定をする
    def game_set(self,x,y):
        if y!=-1:
            self.game_copy(chr(ord('A') + 1+(x-y)))
            self.game_copy(chr(ord('A') + 4+2*x-(x-y)))
        else:
            self.game_copy(chr(ord('A') + 2+x))
    
    #シードの試合をいれる
    def seed_set(self,lit):
        for i in range(1,self.opr.max_row+1):
            name=self.opr.cell(row=i, column=2).value
            num=self.opr.cell(row=i, column=3).value
            if self.opr["B"+str(i)].value==lit:
                if self.tur[chr(ord(name)+1)+str(num-2)].value:
                    self.opr["E"+str(i)]=self.tur[chr(ord(name)+1)+str(num-2)].value
                    self.opr["F"+str(i)]=self.tur[chr(ord(name)+1)+str(num-1)].value
                if self.tur[chr(ord(name)+1)+str(num+2)].value:
                    self.opr["I"+str(i)]=self.tur[chr(ord(name)+1)+str(num+2)].value
                    self.opr["J"+str(i)]=self.tur[chr(ord(name)+1)+str(num+3)].value 
            if self.opr["B"+str(i)].value=='B':
                if self.tur[chr(ord(name)-1)+str(num-2)].value:
                    self.opr["E"+str(i)]=self.tur[chr(ord(name)-1)+str(num-2)].value
                    self.opr["F"+str(i)]=self.tur[chr(ord(name)-1)+str(num-1)].value
                if self.tur[chr(ord(name)-1)+str(num+2)].value:
                    self.opr["I"+str(i)]=self.tur[chr(ord(name)-1)+str(num+2)].value
                    self.opr["J"+str(i)]=self.tur[chr(ord(name)-1)+str(num+3)].value
    
    #シードではない試合を入れる
    def non_seed_set(self,lit):
        for i in range(1,self.opr.max_row+1):
            name=self.opr.cell(row=i, column=2).value
            num=self.opr.cell(row=i, column=3).value
            if self.opr["B"+str(i)].value==lit:
                if self.tur[chr(ord(name)+2)+str(num-4)].value:
                    self.opr["E"+str(i)]=self.tur[chr(ord(name)+2)+str(num-4)].value
                    self.opr["F"+str(i)]=self.tur[chr(ord(name)+2)+str(num-3)].value
                if self.tur[chr(ord(name)+2)+str(num+4)].value:
                    self.opr["I"+str(i)]=self.tur[chr(ord(name)+2)+str(num+4)].value
                    self.opr["J"+str(i)]=self.tur[chr(ord(name)+2)+str(num+5)].value
            if self.opr["B"+str(i)].value=='C':
                if self.tur[chr(ord(name)-2)+str(num-4)].value:
                    self.opr["E"+str(i)]=self.tur[chr(ord(name)-2)+str(num-4)].value
                    self.opr["F"+str(i)]=self.tur[chr(ord(name)-2)+str(num-3)].value
                if self.tur[chr(ord(name)-2)+str(num+4)].value:
                    self.opr["I"+str(i)]=self.tur[chr(ord(name)-2)+str(num+4)].value
                    self.opr["J"+str(i)]=self.tur[chr(ord(name)-2)+str(num+5)].value

    #最初に入る試合を入れておく
    def init_set(self):
        if self.n==5:
            self.non_seed_set('N')
            self.seed_set('O')
        if self.n==4:
            self.non_seed_set('L')
            self.seed_set('M')
        if self.n==3:
            self.non_seed_set('J')
            self.seed_set('K')
        if self.n==2:
            self.non_seed_set('H')
            self.seed_set('I')
        if self.n==1:
            self.non_seed_set('F')
            self.seed_set('G')
    
    #日付メンバー表とトーナメント表から当日使用する管理表の作成
    def setting(self,court_number):
        self.tournament=openpyxl.load_workbook("0527トーナメント.xlsx")
        self.member=openpyxl.load_workbook("0527メンバー.xlsx")
        self.wb = openpyxl.Workbook()
        self.num_list=[]
        for r in range(len(self.tournament.sheetnames)):
            self.mem=self.member.worksheets[r]
            self.tur=self.tournament.worksheets[r]
            self.wb.create_sheet(self.tournament.sheetnames[r])
            self.opr=self.wb.worksheets[r+1]
            self.num=self.mem.max_row-1
            for i in range(10):
                if self.num<=2**i:
                    self.cnt=i
                    break
            self.game_number_search()

            self.opr["A2"].value=self.num
            if  self.num<=8:
                self.num_list.append(1)
            elif self.num<=16:
                self.num_list.append(2)
            elif self.num<=32:
                self.num_list.append(3)
            elif self.num<=64:
                self.num_list.append(4)
            else:
                self.num_list.append(5)
            self.n=self.num_list[r]
            self.init_set()


        self.wb.remove(self.wb.worksheets[0])

        self.wb.create_sheet(index=0, title="順番")
        self.wb.create_sheet(index=1, title="試合番号表")
        self.wb.create_sheet(index=2, title="コート番号表")
        self.rot=self.wb.worksheets[0]
        self.game=self.wb.worksheets[1]
        self.court=self.wb.worksheets[2]
        copy_list = copy.deepcopy(self.num_list)
        for i in range(4,-2,-1):
            for r in range(len(self.tournament.sheetnames)-1,-1,-1):
                self.opr=self.wb.worksheets[r+3]
                self.rank=self.wb.sheetnames[r+3][0:2]
                if self.num_list[r]==i:
                    self.game_set(copy_list[r],self.num_list[r])
                    self.num_list[r]=i-1
        for i in range(1,self.game.max_row+1):
            row=self.rot.max_row+1
            if self.game['E'+str(i)].value!=None and self.game['G'+str(i)].value!=None:
                    self.rot['A'+str(row)]=self.game['A'+str(i)].value
                    self.rot['B'+str(row)]=self.game['D'+str(i)].value
                    self.rot['C'+str(row)]=self.game['E'+str(i)].value
                    self.rot['D'+str(row)]=self.game['F'+str(i)].value
                    self.rot['E'+str(row)]=self.game['G'+str(i)].value
                    self.rot['F'+str(row)]=self.game['H'+str(i)].value

        for i in range(1,court_number+1):
            self.court['A'+str(i)].value=i
            self.court['B'+str(i)]=self.rot['A'+str(2)].value
            self.court['C'+str(i)]=self.rot['B'+str(2)].value
            self.court['D'+str(i)]=self.rot['C'+str(2)].value
            self.court['E'+str(i)]=self.rot['D'+str(2)].value
            self.court['J'+str(i)]=self.rot['E'+str(2)].value
            self.court['K'+str(i)]=self.rot['F'+str(2)].value
            for r in range(1,self.game.max_row+1): 
                if self.rot['B'+str(2)].value==self.game['D'+str(r)].value:
                    self.game['I'+str(r)].value="済"
                    self.book=openpyxl.load_workbook("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(self.court['C'+str(i)].value)+".xlsx")
                    self.score=self.book.worksheets[0]
                    self.score["N4"]=self.court['D'+str(i)].value
                    self.score["B11"]=self.court['D'+str(i)].value
                    self.score["B19"]=self.court['D'+str(i)].value
                    self.score["B27"]=self.court['D'+str(i)].value
                    self.score["AG4"]=self.court['J'+str(i)].value
                    self.score["B13"]=self.court['J'+str(i)].value
                    self.score["B21"]=self.court['J'+str(i)].value
                    self.score["B29"]=self.court['J'+str(i)].value
                    self.score["A4"]=self.court['B'+str(i)].value
                    self.score["E4"]=self.court['C'+str(i)].value
                    self.score["AR4"]=self.court['A'+str(i)].value
                    if str(self.court['E'+str(i)].value)[-2]=='M' or str(self.court['E'+str(i)].value)[-2]=='D':
                        self.score["N8"]=str(self.court['E'+str(i)].value)[1:-3]
                        self.score["V8"]=str(self.court['E'+str(i)].value)[-4:-2]
                    else:
                        self.score["N8"]=str(self.court['E'+str(i)].value)[1:-2]
                        self.score["V8"]=str(self.court['E'+str(i)].value)[-2]
                    if str(self.court['E'+str(i)].value)[-2]=='M' or str(self.court['E'+str(i)].value)[-2]=='D':
                        self.score["AG8"]=str(self.court['K'+str(i)].value)[1:-3]
                        self.score["AO8"]=str(self.court['K'+str(i)].value)[-4:-2]
                    else:
                        self.score["AG8"]=str(self.court['K'+str(i)].value)[1:-2]
                        self.score["AO8"]=str(self.court['K'+str(i)].value)[-2]
                    self.book.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(self.court['C'+str(i)].value)+".xlsx")

            self.rot.delete_rows(2)

        self.tournament.save("トーナメント.xlsx")
        self.wb.save('管理表.xlsx')

    #得点板の初期設定
    def start(self,filename,num):
        wb=openpyxl.load_workbook(filename)
        court=wb.worksheets[2]
        slack = slackweb.Slack(url="https://hooks.slack.com/services/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")
        slack.notify(text="-court"+str(num))
        time.sleep(5)

        for i in range(1,court.max_row+1):
            result=str(court['A'+str(i)].value)+","+court['B'+str(i)].value+","+str(court['C'+str(i)].value)+","+court['D'+str(i)].value+","+court['E'+str(i)].value+","+court['J'+str(i)].value+","+court['K'+str(i)].value
            slack.notify(text="-game"+result)
            time.sleep(5)

if __name__ == '__main__':
    tur=Tournament()
    tur.setting(12)









