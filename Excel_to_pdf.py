#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import win32com.client
from pywintypes import com_error
import pathlib
import os
import pprint

def main(WB_PATH):
    # 絶対パスで指定してね
    p_file = pathlib.Path(WB_PATH)
    PATH_TO_PDF = str(p_file)
    PATH_TO_PDF=PATH_TO_PDF.replace(".xlsx",".pdf")
    print(PATH_TO_PDF)



    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False

    try:
        print('PDFへ変換開始')

        # 開く
        wb = excel.Workbooks.Open(WB_PATH)
        wbk=openpyxl.load_workbook(WB_PATH)
        length=len(wbk.sheetnames)
        ws_index_list=[]
        for i in range(length):
            ws_index_list.append(i+1)
        # 保存したいシートをインデックスで指定。1が最初（一番左）のシート。
        wb.WorkSheets(ws_index_list).Select()

        # 保存
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        print('失敗しました')
    else:
        print('成功しました')
    finally:
        wb.Close()
        excel.Quit()

def main2(file):
    WB_PATH="C:/Users/mech-user/Desktop/IoT/"+file+".xlsx"
    # 絶対パスで指定してね
    p_file = pathlib.Path(WB_PATH)
    PATH_TO_PDF="C:/Users/mech-user/Desktop/IoT/pdf/"+file+".pdf"
    print(PATH_TO_PDF)



    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False

    try:
        print('PDFへ変換開始')

        # 開く
        wb = excel.Workbooks.Open(WB_PATH)
        wbk=openpyxl.load_workbook(WB_PATH)
        length=len(wbk.sheetnames)
        ws_index_list=[]
        wq_index_list=[1,3]
        for i in range(length):
            ws_index_list.append(i+1)
        if file=="管理表":
            wb.WorkSheets(wq_index_list).Select()
        else:
            # 保存したいシートをインデックスで指定。1が最初（一番左）のシート。
            wb.WorkSheets(ws_index_list).Select()

        # 保存
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        print('失敗しました')
    else:
        print('成功しました')
    finally:
        wb.Close()
        excel.Quit()

def date_cate_make(file):
    global wbk
    wbk=openpyxl.load_workbook(file)
    date_list=[]
    date_cate_list=[]
    for i in range(len(wbk.sheetnames)):
        date_list.append(wbk.sheetnames[i][2:])
    date_cate_list=list(set(date_list))
    date_cate_list.sort()
    return date_cate_list

def date_select(file,name):
    global wbk
    wbk=openpyxl.load_workbook(file)
    sheetNames = wbk.get_sheet_names() 
    for i in range(len(sheetNames)):
        if name!=sheetNames[len(sheetNames)-i-1][2:]:
            wbk.remove(wbk.worksheets[len(sheetNames)-i-1])
    wbk.save(name+".xlsx")
    return len(wbk.sheetnames)



        
