#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import numpy as np
from openpyxl import utils
import copy


def date_print(tournament_file,member_file):
    tournament=openpyxl.load_workbook(tournament_file)
    member=openpyxl.load_workbook(member_file)
    date_list=[]
    for r in range(len(tournament.sheetnames)):
        date_list.append(tournament.sheetnames[r][2:6])

    date_list=list(set(date_list))
    date_list.sort()
    return date_list

def date_making(tournament_file,member_file,date):
    tournament=openpyxl.load_workbook(tournament_file)
    member=openpyxl.load_workbook(member_file)

    cnt=0
    for r in range(len(tournament.sheetnames)):
        if tournament.sheetnames[r+cnt][2:6]!=date:
            tournament.remove(tournament.worksheets[r+cnt])
            cnt-=1
        if member.sheetnames[r+cnt][2:6]!=date:
            member.remove(member.worksheets[r+cnt])
    tournament.save(date+'トーナメント.xlsx')
    member.save(date+'メンバー.xlsx')