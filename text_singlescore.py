#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
#0 S  0:0  R 0

def server_selection(server,game):
    if server=="S":
        score['H'+str(11+8*game)]='S'
    else:
        score['H'+str(13+8*game)]='S'
    score['I'+str(11+8*game)]=0
    score['I'+str(13+8*game)]=0
    return 0
def game_end(player_point,enemy_point,game):
    if player_point>=21 and player_point-enemy_point>=2:
        score['Z'+str(4+2*game)]=player_point
        score['AC'+str(4+2*game)]=enemy_point
    elif enemy_point>=21 and enemy_point-player_point>=2:
        score['Z'+str(4+2*game)]=player_point
        score['AC'+str(4+2*game)]=enemy_point


def point_write(player_point,enemy_point,game,max_column):
    i=9+player_point+enemy_point
    print("-")
    print(player_point)
    print(enemy_point)
    print(i)
    print("-")

    if i<=max_column:
        if player_point!=score.cell(row=11+8*game, column=i-1).value:
            score.cell(row=11+8*game, column=i).value=player_point
        else:
            score.cell(row=13+8*game, column=i).value=enemy_point
    else:
        if player_point!=score.cell(row=15+8*game, column=i-max_column+7).value:
            score.cell(row=15+8*game, column=i-max_column+8).value=player_point
        else:
            score.cell(row=17+8*game, column=i-max_column+8).value=enemy_point


def read_text_score(wb):
    global score
    f = open("C:\\Users\\mech-user\\Desktop\\IoT\\試合番号データ\\"+str(wb)+".txt", 'r')
    scorebook = openpyxl.load_workbook("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(wb)+".xlsx",read_only=False)
    score = scorebook.worksheets[0]
    max_column=score.max_column-2
    print(max_column)
    line = f.readline()
    while line:
        line=line.strip()
        info=line.split(" ")
        player_serve=info[1]
        player_gamenum=int(info[0])
        point=info[2].split(":")
        player_point=int(point[0])
        enemy_point=int(point[1])
        enemy_gamenum=int(info[4])
        game=player_gamenum+enemy_gamenum
        if player_point==0 and enemy_point==0 and player_gamenum!=2 and enemy_gamenum!=2:
            server_selection(player_serve,game)
        point_write(player_point,enemy_point,game,max_column)
        game_end(player_point,enemy_point,game)
        line = f.readline()
    score["X4"]=player_gamenum
    score["AE4"]=enemy_gamenum
    f.close()
    print("書き込み完了")
    scorebook.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(wb)+".xlsx")


