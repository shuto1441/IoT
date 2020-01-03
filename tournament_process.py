#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import numpy as np
from openpyxl import utils
import copy
import slackweb
import time
import Excel_to_png
import post_tweet
import ftp_upload
import os
import client
import shutil
import text_singlescore
from plyer import notification
import random

right = Border(right=Side(style='thick', color='FF0000'))
left = Border(left=Side(style='thick', color='FF0000'))
top_right = Border(top=Side(style='thick', color='FF0000'),
                right=Side(style='thick', color='FF0000')
)
under_right = Border(bottom=Side(style='thick', color='FF0000'),
                right=Side(style='thick', color='FF0000')
)
top_left = Border(top=Side(style='thick', color='FF0000'),
                left=Side(style='thick', color='FF0000')
)
under_left_right = Border(bottom=Side(style='thick', color='FF0000'),
                left=Side(style='thick', color='FF0000'),
                right=Side(style='thick', color='FF0000')
)
under_left = Border(bottom=Side(style='thick', color='FF0000'),
                left=Side(style='thick', color='FF0000')
)
under=Border(bottom=Side(style='thick', color='FF0000'))
right_center=Alignment(horizontal = 'right',
                        vertical = 'center')
left_center=Alignment(horizontal = 'left',
                        vertical = 'center')
under_black_right=Border(bottom=Side(style='thin', color='000000'),
            right=Side(style='thick', color='FF0000'))
under_black_left=Border(bottom=Side(style='thin', color='000000'),
            left=Side(style='thick', color='FF0000'))
right_black = Border(right=Side(style='thin', color='000000'))
left_black = Border(left=Side(style='thin', color='000000'))

def red_line_draw_right(self,result,num,cnt,player,enemy):
    if result==1:
        tur.unmerge_cells(self+str(num)+':'+self+str(num+1))
        tur[self+str(num-1-cnt)].border=top_right
        tur[chr(ord(self)+1)+str(num)].border=under_left
        tur[self+str(num-2-cnt)]=player
        tur[self+str(num-2-cnt)].alignment = right_center
        tur[self+str(num+3+cnt)]=enemy
        tur[self+str(num+3+cnt)].alignment = right_center
        if cnt>0:
            for i in range(cnt):
                tur[self+str(num-1-i)].border=right
            tur[chr(ord(self)+1)+str(num+cnt+2)].border=left_black
        tur.merge_cells(self+str(num)+':'+self+str(num+1))

    elif result==-1:
        tur.unmerge_cells(self+str(num)+':'+self+str(num+1))
        tur[self+str(num+2+cnt)].border=under_right
        tur[chr(ord(self)+1)+str(num+1)].border=top_left
        tur[self+str(num-2-cnt)]=player
        tur[self+str(num-2-cnt)].alignment = right_center
        tur[self+str(num+3+cnt)]=enemy
        tur[self+str(num+3+cnt)].alignment = right_center
        if cnt>0:
            for i in range(cnt):
                tur[self+str(num+2+i)].border=right
            tur[chr(ord(self)+1)+str(num-cnt)].border=left_black
        tur.merge_cells(self+str(num)+':'+self+str(num+1))

def red_line_draw_left(self,result,num,cnt,player,enemy):
    if result==1:
        tur.unmerge_cells(self+str(num)+':'+self+str(num+1))
        tur[self+str(num-1-cnt)].border=top_left
        tur[chr(ord(self)-1)+str(num)].border=under_right
        tur[self+str(num-2-cnt)]=player
        tur[self+str(num-2-cnt)].alignment = left_center
        tur[self+str(num+3+cnt)]=enemy
        tur[self+str(num+3+cnt)].alignment = left_center
        if cnt>0:
            for i in range(cnt):
                tur[self+str(num-1-i)].border=left
            tur[chr(ord(self)-1)+str(num+cnt+2)].border=right_black
        tur.merge_cells(self+str(num)+':'+self+str(num+1))

    elif result==-1:
        tur.unmerge_cells(self+str(num)+':'+self+str(num+1))
        tur[self+str(num+2+cnt)].border=under_left
        tur[chr(ord(self)-1)+str(num+1)].border=top_right
        tur[self+str(num-2-cnt)]=player
        tur[self+str(num-2-cnt)].alignment = left_center
        tur[self+str(num+3+cnt)]=enemy
        tur[self+str(num+3+cnt)].alignment = left_center
        if cnt>0:
            for i in range(cnt):
                tur[self+str(num+2+i)].border=left
            tur[chr(ord(self)-1)+str(num-cnt)].border=right_black
        tur.merge_cells(self+str(num)+':'+self+str(num+1))

def red_line_draw_final(self,result,num,player,enemy):
    if result==1:
        tur.unmerge_cells(self+str(num)+':'+chr(ord(self)+1)+str(num))
        tur[self+str(num-1)].border=under_left_right
        tur[chr(ord(self)-1)+str(num+1)]=player
        tur[chr(ord(self)-1)+str(num+1)].alignment = right_center
        tur[chr(ord(self)+2)+str(num+1)]=enemy
        tur[chr(ord(self)+2)+str(num+1)].alignment = left_center
        tur[self+str(num-2)].border=right
        tur.merge_cells(self+str(num)+':'+chr(ord(self)+1)+str(num))

    elif result==-1:
        tur.unmerge_cells(self+str(num)+':'+chr(ord(self)+1)+str(num))
        tur[chr(ord(self)+1)+str(num-1)].border=under_left_right
        tur[chr(ord(self)-1)+str(num+1)]=player
        tur[chr(ord(self)-1)+str(num+1)].alignment = right_center
        tur[chr(ord(self)+2)+str(num+1)]=enemy
        tur[chr(ord(self)+2)+str(num+1)].alignment = left_center
        tur[chr(ord(self)+1)+str(num-2)].border=left
        tur.merge_cells(self+str(num)+':'+chr(ord(self)+1)+str(num))


def home_left_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)+1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["E"+str(i)].value
            opr["F"+str(j)]=opr["F"+str(i)].value
            game_left_search(j)
        if opr["B"+str(j)].value==chr(ord(name)+1) and opr["C"+str(j)].value==opr["C"+str(i)].value-cnt:
            opr["I"+str(j)]=opr["E"+str(i)].value
            opr["J"+str(j)]=opr["F"+str(i)].value
            game_right_search(j)

def home_right_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)-1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["E"+str(i)].value
            opr["F"+str(j)]=opr["F"+str(i)].value
            game_left_search(j)
        if opr["B"+str(j)].value==chr(ord(name)-1) and opr["C"+str(j)].value==opr["C"+str(i)].value-cnt:
            opr["I"+str(j)]=opr["E"+str(i)].value
            opr["J"+str(j)]=opr["F"+str(i)].value
            game_right_search(j)

def home_center_left_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)+1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["E"+str(i)].value
            opr["F"+str(j)]=opr["F"+str(i)].value
            game_left_search(j)

def away_center_left_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)+1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["I"+str(i)].value
            opr["F"+str(j)]=opr["J"+str(i)].value
            game_left_search(j)
def home_center_right_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)-2)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["I"+str(j)]=opr["E"+str(i)].value
            opr["J"+str(j)]=opr["F"+str(i)].value
            game_right_search(j)

def away_center_right_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)-2)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["I"+str(j)]=opr["I"+str(i)].value
            opr["J"+str(j)]=opr["J"+str(i)].value
            game_right_search(j)

def away_left_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)+1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["I"+str(i)].value
            opr["F"+str(j)]=opr["J"+str(i)].value
            game_left_search(j)
        if opr["B"+str(j)].value==chr(ord(name)+1) and opr["C"+str(j)].value==opr["C"+str(i)].value-cnt:
            opr["I"+str(j)]=opr["I"+str(i)].value
            opr["J"+str(j)]=opr["J"+str(i)].value
            game_right_search(j)

def away_right_serch(name,i,cnt):
    for j in range(1,opr.max_row+1):
        if opr["B"+str(j)].value==chr(ord(name)-1)and opr["C"+str(j)].value==opr["C"+str(i)].value+cnt:
            opr["E"+str(j)]=opr["I"+str(i)].value
            opr["F"+str(j)]=opr["J"+str(i)].value
            game_left_search(j)
        if opr["B"+str(j)].value==chr(ord(name)-1) and opr["C"+str(j)].value==opr["C"+str(i)].value-cnt:
            opr["I"+str(j)]=opr["I"+str(i)].value
            opr["J"+str(j)]=opr["J"+str(i)].value
            game_right_search(j)

def game_left_search(j):
    for k in range(1,game.max_row+1):
        if game["D"+str(k)].value==opr["D"+str(j)].value:
            game["E"+str(k)].value=opr["E"+str(j)].value
            game["F"+str(k)].value=opr["F"+str(j)].value

def game_right_search(j):
    for k in range(1,game.max_row+1):
        if game["D"+str(k)].value==opr["D"+str(j)].value:
            game["G"+str(k)].value=opr["I"+str(j)].value
            game["H"+str(k)].value=opr["J"+str(j)].value


def add_text(game,cnt,tex):
    filename="C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(game)+"-"+str(cnt)+".txt"
    with open(filename,'a') as file_object:
        file_object.write(tex+'\n')
def filemake(game,cnt):
    filename="C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(game)+"-"+str(cnt)+".txt"
    with open(filename,'w'):
        pass

def update_tournament(num,n):
    for i in range(1,opr.max_row+1):
        if num<=8:
            if opr["G"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    home_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    home_center_left_serch('C',i,1)
                if opr["B"+str(i)].value=='G':
                    home_right_serch('G',i,4)
                if opr["B"+str(i)].value=='F':
                    home_center_right_serch('F',i,1)
            if opr["H"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    away_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    away_center_left_serch('C',i,1)
                if opr["B"+str(i)].value=='G':
                    away_right_serch('G',i,4)
                if opr["B"+str(i)].value=='F':
                    away_center_right_serch('F',i,1)
        elif num<=16:
            if opr["G"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    home_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    home_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    home_center_left_serch('D',i,1)
                if opr["B"+str(i)].value=='I':
                    home_right_serch('I',i,4)
                if opr["B"+str(i)].value=='H':
                    home_right_serch('H',i,8)
                if opr["B"+str(i)].value=='G':
                    home_center_right_serch('G',i,1)
            if opr["H"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    away_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    away_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    away_center_left_serch('D',i,1)
                if opr["B"+str(i)].value=='I':
                    away_right_serch('I',i,4)
                if opr["B"+str(i)].value=='H':
                    away_right_serch('H',i,8)
                if opr["B"+str(i)].value=='G':
                    away_center_right_serch('G',i,1)
        elif num<=32:
            if opr["G"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    home_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    home_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    home_left_serch('D',i,16)
                if opr["B"+str(i)].value=='E':
                    home_center_left_serch('E',i,1)
                if opr["B"+str(i)].value=='K':
                    home_right_serch('K',i,4)
                if opr["B"+str(i)].value=='J':
                    home_right_serch('J',i,8)
                if opr["B"+str(i)].value=='I':
                    home_right_serch('I',i,16)
                if opr["B"+str(i)].value=='H':
                    home_center_right_serch('H',i,1)
            if opr["H"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    away_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    away_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    away_left_serch('D',i,16)
                if opr["B"+str(i)].value=='E':
                    away_center_left_serch('E',i,1)
                if opr["B"+str(i)].value=='K':
                    away_right_serch('K',i,4)
                if opr["B"+str(i)].value=='J':
                    away_right_serch('J',i,8)
                if opr["B"+str(i)].value=='I':
                    away_right_serch('I',i,16)
                if opr["B"+str(i)].value=='H':
                    away_center_right_serch('H',i,1)
        else:
            if opr["G"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    home_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    home_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    home_left_serch('D',i,16)
                if opr["B"+str(i)].value=='E':
                    home_left_serch('E',i,32)
                if opr["B"+str(i)].value=='F':
                    home_center_left_serch('F',i,1)
                if opr["B"+str(i)].value=='M':
                    home_right_serch('M',i,4)
                if opr["B"+str(i)].value=='L':
                    home_right_serch('L',i,8)
                if opr["B"+str(i)].value=='K':
                    home_right_serch('K',i,16)
                if opr["B"+str(i)].value=='J':
                    home_right_serch('J',i,32)
                if opr["B"+str(i)].value=='I':
                    home_center_right_serch('I',i,1)
            if opr["H"+str(i)].value==2:
                if opr["B"+str(i)].value=='B':
                    away_left_serch('B',i,4)
                if opr["B"+str(i)].value=='C':
                    away_left_serch('C',i,8)
                if opr["B"+str(i)].value=='D':
                    away_left_serch('D',i,16)
                if opr["B"+str(i)].value=='E':
                    away_left_serch('E',i,32)
                if opr["B"+str(i)].value=='F':
                    away_center_left_serch('F',i,1)
                if opr["B"+str(i)].value=='M':
                    away_right_serch('M',i,4)
                if opr["B"+str(i)].value=='L':
                    away_right_serch('L',i,8)
                if opr["B"+str(i)].value=='K':
                    away_right_serch('K',i,16)
                if opr["B"+str(i)].value=='J':
                    away_right_serch('J',i,32)
                if opr["B"+str(i)].value=='I':
                    away_center_right_serch('I',i,1)

    for i in range(1,opr.max_row+1):
        result=0
        player=opr["G"+str(i)].value
        enemy=opr["H"+str(i)].value
        if opr["G"+str(i)].value==2:
            result=1
        elif opr["H"+str(i)].value==2:
            result=-1
        if result!=0:
            if i==opr.max_row:
                red_line_draw_final(opr["B"+str(i)].value,result,opr["C"+str(i)].value,player,enemy)
            else:
                if opr["B"+str(i)].value=='B':
                    red_line_draw_right('B',result,opr["C"+str(i)].value,0,player,enemy)
                elif opr["B"+str(i)].value=='C':
                    red_line_draw_right('C',result,opr["C"+str(i)].value,2,player,enemy)
                    if n==8:
                        tur["D12"].border=under_black_left
                elif opr["B"+str(i)].value=='D' and n>8:
                    red_line_draw_right('D',result,opr["C"+str(i)].value,6,player,enemy)
                    if n==16:
                        tur["E20"].border=under_black_left
                elif opr["B"+str(i)].value=='E' and n>16:
                    red_line_draw_right('E',result,opr["C"+str(i)].value,14,player,enemy)
                    if n==32:
                        tur["F36"].border=under_black_left
                elif opr["B"+str(i)].value=='F' and n>32:
                    red_line_draw_right('F',result,opr["C"+str(i)].value,30,player,enemy)
                    tur["G68"].border=under_black_left
                elif opr["B"+str(i)].value=='F' and n==8:
                    red_line_draw_left('F',result,opr["C"+str(i)].value,2,player,enemy)
                    tur["E12"].border=under_black_right
                elif opr["B"+str(i)].value=='G' and n==8:
                    red_line_draw_left('G',result,opr["C"+str(i)].value,0,player,enemy)
                elif opr["B"+str(i)].value=='G' and n==16:
                    red_line_draw_left('G',result,opr["C"+str(i)].value,6,player,enemy)
                    tur["F20"].border=under_black_right
                elif opr["B"+str(i)].value=='H' and n==16:
                    red_line_draw_left('H',result,opr["C"+str(i)].value,2,player,enemy)
                elif opr["B"+str(i)].value=='I' and n==16:
                    red_line_draw_left('I',result,opr["C"+str(i)].value,0,player,enemy)
                elif opr["B"+str(i)].value=='H' and n==32:
                    red_line_draw_left('H',result,opr["C"+str(i)].value,14,player,enemy)
                    tur["G36"].border=under_black_right
                elif opr["B"+str(i)].value=='I' and n==32:
                    red_line_draw_left('I',result,opr["C"+str(i)].value,6,player,enemy)
                elif opr["B"+str(i)].value=='J' and n==32:
                    red_line_draw_left('J',result,opr["C"+str(i)].value,2,player,enemy)
                elif opr["B"+str(i)].value=='K' and n==32:
                    red_line_draw_left('K',result,opr["C"+str(i)].value,0,player,enemy)
                elif opr["B"+str(i)].value=='I' and n==64:
                    red_line_draw_left('I',result,opr["C"+str(i)].value,30,player,enemy)
                    tur["H68"].border=under_black_right
                elif opr["B"+str(i)].value=='J' and n==64:
                    red_line_draw_left('J',result,opr["C"+str(i)].value,14,player,enemy)
                elif opr["B"+str(i)].value=='K' and n==64:
                    red_line_draw_left('K',result,opr["C"+str(i)].value,6,player,enemy)
                elif opr["B"+str(i)].value=='L' and n==64:
                    red_line_draw_left('L',result,opr["C"+str(i)].value,2,player,enemy)
                elif opr["B"+str(i)].value=='M' and n==64:
                    red_line_draw_left('M',result,opr["C"+str(i)].value,0,player,enemy)

def read_text(file):
    f=open("C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(file)+".txt")
    line=f.readline()
    ans='0 S 0:0 R 0'
    while line:
        ans=line
        line=f.readline()			
    f.close()
    ans=ans.strip()
    result=ans.split()
    left=int(result[0])
    right=int(result[4])
    print(result)
    point=result[2]
    result2=point.split(":")
    left2=int(result2[0])
    right2=int(result2[1])
    return left,left2,right2,right

def notify(text):
    notification.notify(
        title='試合情報をお知らせします',
        message=text,
        app_name='試合通知',
        app_icon='./バドミントン.ico'
    )




def main(file1,file2):
    global opr,tur,player,enemy,player_game,enemy_game,serve,pre_player,pre_enemy,serve2,game
    cnt=0
    precnt=0
    wb=openpyxl.load_workbook(file1)
    court=wb.worksheets[2]
    slack = slackweb.Slack(url="https://hooks.slack.com/services/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")
    
    """
    if os.path.exists('C:\\Users\\mech-user\\Desktop\\IoT\\pdf'):
        shutil.rmtree("./pdf")
    os.mkdir("./pdf")
    if os.path.exists('C:\\Users\\mech-user\\Desktop\\IoT\\pi'):
        shutil.rmtree("./pi")
    os.mkdir("./pi")
    if os.path.exists('C:\\Users\\mech-user\\Desktop\\IoT\\試合番号データ'):
        shutil.rmtree("./試合番号データ")
    os.mkdir("./試合番号データ")
    """

    player=[0]*(court.max_row-1)
    enemy=[0]*(court.max_row-1)
    pre_player=[0]*(court.max_row-1)
    pre_enemy=[0]*(court.max_row-1)
    player_game=[0]*(court.max_row-1)
    enemy_game=[0]*(court.max_row-1)
    serve=['S']*(court.max_row-1)
    serve2=['R']*(court.max_row-1)
    count=[1000]*(court.max_row-1)
    for i in range(1,court.max_row+1):
        if i!=court.max_row:
            filemake(str(court['C'+str(i)].value),count[i-1])
            text="0 S 0:0 R 0"
            add_text(str(court['C'+str(i+1)].value),count[i-1],text)
    while True:
        wb=openpyxl.load_workbook(file1)
        tournament=openpyxl.load_workbook(file2)
        print("ok")
        rot=wb.worksheets[0]
        game=wb.worksheets[1]
        court=wb.worksheets[2]
        for r in range(len(wb.sheetnames)-3): 
            opr=wb.worksheets[r+3]
            tur=tournament.worksheets[r]
            if opr["A2"].value<=8:
                n=8
            elif opr["A2"].value<=16:
                n=16
            elif opr["A2"].value<=32:
                n=32
            else:
                n=64
            update_tournament(opr["A2"].value,n)


        #順番表の更新
        rot.delete_rows(idx=2, amount=rot.max_row)
        for i in range(1,game.max_row+1):
            row=rot.max_row+1
            if game['E'+str(i)].value!=None and game['G'+str(i)].value!=None and game['I'+str(i)].value==None:
                    rot['A'+str(row)].value=game['A'+str(i)].value
                    rot['B'+str(row)].value=game['D'+str(i)].value
                    rot['C'+str(row)].value=game['E'+str(i)].value
                    rot['D'+str(row)].value=game['F'+str(i)].value
                    rot['E'+str(row)].value=game['G'+str(i)].value
                    rot['F'+str(row)].value=game['H'+str(i)].value

        client.client = client._blob_client()
        local_ =  client.client.fetch_local()
        remote_ =  client.client.fetch_remote()
        download_blobs = []
        remote_blobs=[]
        remove_blobs=[]

        #コート番号表の更新
        for i in range(1,court.max_row+1):
            if court['L'+str(i)].value=="終了":
                cnt+=1
                if i!=court.max_row:
                    player[i-1]=0
                    enemy[i-1]=0
                    pre_player[i-1]=0
                    pre_enemy[i-1]=0
                    player_game[i-1]=0
                    enemy_game[i-1]=0
                    serve[i-1]='S'
                    serve2[i-1]='R'
                if i==court.max_row:
                    slack.notify(text="-scb off")
                    time.sleep(3)
                for item_local in local_:
                    if str(court['C'+str(i)].value)==str(item_local)[1:4]:
                        os.rename('C:\\Users\\mech-user\\Desktop\\IoT\\pi\\'+item_local[1:],'C:\\Users\\mech-user\\Desktop\\IoT\\pi\\'+item_local[1:4]+'.txt')
                        new_path = shutil.move('C:\\Users\\mech-user\\Desktop\\IoT\\pi\\'+item_local[1:4]+'.txt', "C:\\Users\\mech-user\\Desktop\\IoT\\試合番号データ\\")
                        text_singlescore.read_text_score(court['C'+str(i)].value)

                for r in range(3,len(wb.sheetnames)):
                    opr=wb.worksheets[r]
                    rank=wb.sheetnames[r][0:2]  
                    if rank==court['B'+str(i)].value:
                        for j in range(1,opr.max_row+1):
                            if court['C'+str(i)].value==opr['D'+str(j)].value:
                                opr['G'+str(j)].value=court['F'+str(i)].value
                                opr['H'+str(j)].value=court['I'+str(i)].value
                                court['L'+str(i)].value=""
                                court['F'+str(i)].value=0
                                court['G'+str(i)].value=0
                                court['H'+str(i)].value=0
                                court['I'+str(i)].value=0
                for r in range(1,game.max_row+1): 
                    if rot['B'+str(2)].value==game['D'+str(r)].value:
                        game['I'+str(r)].value="済"
                court['B'+str(i)].value=rot['A'+str(2)].value
                court['C'+str(i)].value=rot['B'+str(2)].value
                court['D'+str(i)].value=rot['C'+str(2)].value
                court['E'+str(i)].value=rot['D'+str(2)].value
                court['J'+str(i)].value=rot['E'+str(2)].value
                court['K'+str(i)].value=rot['F'+str(2)].value
                result=str(court['A'+str(i)].value)+","+court['B'+str(i)].value+","+str(court['C'+str(i)].value)+","+court['D'+str(i)].value+","+court['E'+str(i)].value+","+court['J'+str(i)].value+","+court['K'+str(i)].value
                slack.notify(text=result)
                notify(result)
                if i!=court.max_row:
                    count[i-1]=1000
                    filemake(str(court['C'+str(i)].value),count[i-1])
                    scorebook=openpyxl.load_workbook("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(court['C'+str(i)].value)+".xlsx")
                    score=scorebook.worksheets[0]
                    score["N4"]=court['D'+str(i)].value
                    score["B11"]=court['D'+str(i)].value
                    score["B19"]=court['D'+str(i)].value
                    score["B27"]=court['D'+str(i)].value
                    score["AG4"]=court['J'+str(i)].value
                    score["B13"]=court['J'+str(i)].value
                    score["B21"]=court['J'+str(i)].value
                    score["B29"]=court['J'+str(i)].value
                    score["A4"]=court['B'+str(i)].value
                    score["E4"]=court['C'+str(i)].value
                    score["AR4"]=court['A'+str(i)].value
                    if str(court['E'+str(i)].value)[-2]=='M' or str(court['E'+str(i)].value)[-2]=='D':
                        score["N8"]=str(court['E'+str(i)].value)[1:-3]
                        score["V8"]=str(court['E'+str(i)].value)[-4:-2]
                    else:
                        score["N8"]=str(court['E'+str(i)].value)[1:-2]
                        score["V8"]=str(court['E'+str(i)].value)[-2]
                    if str(court['E'+str(i)].value)[-2]=='M' or str(court['E'+str(i)].value)[-2]=='D':
                        score["AG8"]=str(court['K'+str(i)].value)[1:-3]
                        score["AO8"]=str(court['K'+str(i)].value)[-4:-2]
                    else:
                        score["AG8"]=str(court['K'+str(i)].value)[1:-2]
                        score["AO8"]=str(court['K'+str(i)].value)[-2]
                    scorebook.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(court['C'+str(i)].value)+".xlsx")
                
                time.sleep(3)
                rot.delete_rows(2)
                if i==court.max_row:
                    slack.notify(text="-scb on")
                    time.sleep(3)
                Excel_to_png.excel_to_png('トーナメント')
                for i in range(len(tournament.sheetnames)):
                    ftp_upload.main('tournament{}'.format(tournament.sheetnames[i][0:2]))


        #Azureの整理
        for item_remote in remote_:
            remote_blobs.append(item_remote)
        remote_blobs.sort()
        print(remote_blobs)
        #Azureのクラウド上から削除するリストの作成
        for i in range(len(remote_blobs)-1,0,-1):
            if remote_blobs[i][3:6]==remote_blobs[i-1][3:6]:
                remove_blobs.append(remote_blobs[i-1])
        print(remove_blobs)
        #Azureから削除する
        for blob in remove_blobs:
            print("remove: {}".format(blob))
            client.client.clear(blob)
        remote_ =  client.client.fetch_remote()
        #Azureからのダウンロードリスト作成
        for item_remote in remote_:
            for i in range(1,court.max_row+1):
                if str(court['C'+str(i)].value)==str(item_remote)[3:6]:
                    download_blobs.append(item_remote)
        for blob in download_blobs:
            for item_local in local_:
                if str(item_local)[1:]==str(blob)[3:]:
                    download_blobs.remove(blob)
        print("local ----")
        for item in local_:
            print(item)
        #Azureからのダウンロード前にローカルファイルの削除
        for item_local in local_:
            #for i in range(1,court.max_row+1):
                #if str(court['C'+str(i)].value)==str(item_local)[1:4] and len(download_blobs)>0:
            if str(court['C'+str(court.max_row)].value)==str(item_local)[1:4] and len(download_blobs)>0:
                print("C:\\Users\\mech-user\\Desktop\\IoT\\pi"+item_local)
                os.remove("C:\\Users\\mech-user\\Desktop\\IoT\\pi"+item_local)
                time.sleep(3)
        #Azureからのダウンロード
        for blob in download_blobs:
            print("download: {}".format(blob))
            client.client.download(blob)

        for i in range(court.max_row-1):
                if player[i]>=21 and player[i]-enemy[i]>=2:
                    player[i]=0
                    enemy[i]=0
                    player_game[i]+=1
                    serve[i]='R'
                    serve2[i]='S'
                elif enemy[i]>=21 and enemy[i]-player[i]>=2:
                    enemy_game[i]+=1
                    player[i]=0
                    enemy[i]=0
                    serve[i]='R'
                    serve2[i]='S'
                elif random.randint(1,2)==1:
                    player[i]+=1
                    if serve[i]=='R':
                        serve[i]='S'
                        serve2[i]='R'
                else:
                    enemy[i]+=1
                    if serve[i]=='S':
                        serve[i]='R'
                        serve2[i]='S'
                text=str(player_game[i])+" "+serve[i]+" "+str(player[i])+":"+str(enemy[i])+" "+serve2[i]+" "+str(enemy_game[i])
                count[i]+=1
                os.rename("C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(court['C'+str(i+1)].value)+"-"+str(count[i]-1)+".txt","C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(court['C'+str(i+1)].value)+"-"+str(count[i])+".txt")
                add_text(str(court['C'+str(i+1)].value),count[i],text)
        
        
        #試合情報のダウンロード後の処理
        for i in range(1,court.max_row+1):
            for j in range(1000,1200):
                if os.path.exists("C:\\Users\\mech-user\\Desktop\\IoT\\pi\\"+str(court['C'+str(i)].value)+"-"+str(j)+".txt"):
                    court['F'+str(i)].value,court['G'+str(i)].value,court['H'+str(i)].value,court['I'+str(i)].value=read_text(str(court['C'+str(i)].value)+"-"+str(j))
                    break
            if court['F'+str(i)].value==2 or court['I'+str(i)].value==2:
                court['L'+str(i)].value="終了"

        wb.save('管理表.xlsx')
        #shutil.rmtree("./pdf")
        #os.mkdir("./pdf")
        #画像変換
        Excel_to_png.excel_to_png('管理表')
        #ツイート投稿
        if int(cnt/3)>int(precnt/3) and cnt>3:
            post_tweet.main('court')
        if int(cnt/5)>int(precnt/5) and cnt>5:
            for i in range(len(tournament.sheetnames)):
                post_tweet.main('tournament{}'.format(tournament.sheetnames[i][0:2]))
        tournament.save('トーナメント.xlsx')
        precnt=cnt
        #サーバーへのアップロード
        ftp_upload.main('court')
        ftp_upload.main('next')

if __name__ == '__main__':
    main("管理表.xlsx","トーナメント.xlsx")
    ledprint(u"このような形で表示をすることが可能です")

