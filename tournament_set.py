#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import numpy as np
from openpyxl import utils
import copy
import re
import slackweb

def list_make(self):
    global tur,opr,book3,score
    for i in range(6,tur.max_row+1):
        if tur[self+str(i)].value:
            if int(tur[self+str(i)].value)>=100:
                row=opr.max_row+1
                opr["B"+str(row)].value=self
                opr["C"+str(row)].value=i
                opr["D"+str(row)].value=tur[self+str(i)].value
                new =openpyxl.load_workbook("バドミントンスコアシート.xlsx")
                score = new['シングルス用スコア']
                new.remove(new['ダブルス用スコア'])
                score.title=str(opr["D"+str(row)].value)
                new.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(opr["D"+str(row)].value)+".xlsx")
                

def game_number_search(num):
    if num<=8:
        list_make('B')
        list_make('G')
        list_make('C')
        list_make('F')
        list_make('D')
    elif num<=16:
        list_make('B')
        list_make('I')
        list_make('C')
        list_make('H')
        list_make('D')
        list_make('G')
        list_make('E')
    elif num<=32:
        list_make('B')
        list_make('K')
        list_make('C')
        list_make('J')
        list_make('D')
        list_make('I')
        list_make('E')
        list_make('H')
        list_make('F')
    else:
        list_make('B')
        list_make('M')
        list_make('C')
        list_make('L')
        list_make('D')
        list_make('K')
        list_make('E')
        list_make('J')
        list_make('F')
        list_make('I')
        list_make('G')

def game_copy(name,rank):
    global tur,opr,game
    for i in range(1,opr.max_row+1):
        row=game.max_row+1
        if opr["B"+str(i)].value==name:
            game["A"+str(row)]=rank
            game["B"+str(row)]=name
            game["C"+str(row)]=opr["C"+str(i)].value
            game["D"+str(row)]=opr["D"+str(i)].value
            if opr["E"+str(i)].value:
                game["E"+str(row)]=opr["E"+str(i)].value
                game["F"+str(row)]=opr["F"+str(i)].value
            if opr["I"+str(i)].value:
                game["G"+str(row)]=opr["I"+str(i)].value
                game["H"+str(row)]=opr["J"+str(i)].value


def game_set(x,y,rank):
    global tur,opr
    if y!=-1:
        game_copy(chr(ord('A') + 1+(x-y)),rank)
        game_copy(chr(ord('A') + 4+2*x-(x-y)),rank)
    else:
        game_copy(chr(ord('A') + 2+x),rank)

def seed_set(lit):
    global tur,opr
    for i in range(1,opr.max_row+1):
        name=opr.cell(row=i, column=2).value
        num=opr.cell(row=i, column=3).value
        if opr["B"+str(i)].value==lit:
            if tur[chr(ord(name)+1)+str(num-2)].value:
                opr["E"+str(i)]=tur[chr(ord(name)+1)+str(num-2)].value
                opr["F"+str(i)]=tur[chr(ord(name)+1)+str(num-1)].value
            if tur[chr(ord(name)+1)+str(num+2)].value:
                opr["I"+str(i)]=tur[chr(ord(name)+1)+str(num+2)].value
                opr["J"+str(i)]=tur[chr(ord(name)+1)+str(num+3)].value 
        if opr["B"+str(i)].value=='B':
            if tur[chr(ord(name)-1)+str(num-2)].value:
                opr["E"+str(i)]=tur[chr(ord(name)-1)+str(num-2)].value
                opr["F"+str(i)]=tur[chr(ord(name)-1)+str(num-1)].value
            if tur[chr(ord(name)-1)+str(num+2)].value:
                opr["I"+str(i)]=tur[chr(ord(name)-1)+str(num+2)].value
                opr["J"+str(i)]=tur[chr(ord(name)-1)+str(num+3)].value
def non_seed_set(lit):
    global tur,opr
    for i in range(1,opr.max_row+1):
            name=opr.cell(row=i, column=2).value
            num=opr.cell(row=i, column=3).value
            if opr["B"+str(i)].value==lit:
                if tur[chr(ord(name)+2)+str(num-4)].value:
                    opr["E"+str(i)]=tur[chr(ord(name)+2)+str(num-4)].value
                    opr["F"+str(i)]=tur[chr(ord(name)+2)+str(num-3)].value
                if tur[chr(ord(name)+2)+str(num+4)].value:
                    opr["I"+str(i)]=tur[chr(ord(name)+2)+str(num+4)].value
                    opr["J"+str(i)]=tur[chr(ord(name)+2)+str(num+5)].value
            if opr["B"+str(i)].value=='C':
                if tur[chr(ord(name)-2)+str(num-4)].value:
                    opr["E"+str(i)]=tur[chr(ord(name)-2)+str(num-4)].value
                    opr["F"+str(i)]=tur[chr(ord(name)-2)+str(num-3)].value
                if tur[chr(ord(name)-2)+str(num+4)].value:
                    opr["I"+str(i)]=tur[chr(ord(name)-2)+str(num+4)].value
                    opr["J"+str(i)]=tur[chr(ord(name)-2)+str(num+5)].value

def init_set(n):
    if n==4:
        non_seed_set('L')
        seed_set('M')
    if n==3:
        non_seed_set('J')
        seed_set('K')
    if n==2:
        non_seed_set('H')
        seed_set('I')
    if n==1:
        non_seed_set('F')
        seed_set('G')


def main(tournament_file,member_file,court_number):
    global tur,opr,sscore,game
    tournament=openpyxl.load_workbook(tournament_file)
    member=openpyxl.load_workbook(member_file)
    wb = openpyxl.Workbook()
    num_list=[]
    for r in range(len(tournament.sheetnames)):
        mem=member.worksheets[r]
        tur=tournament.worksheets[r]
        wb.create_sheet(tournament.sheetnames[r])
        opr=wb.worksheets[r+1]
        num=mem.max_row-1
        game_number_search(num)
        opr["A2"].value=num
        if num<=8:
            num_list.append(1)
        elif num<=16:
            num_list.append(2)
        elif num<=32:
            num_list.append(3)
        else:
            num_list.append(4)

    wb.remove(wb.worksheets[0])

    for r in range(len(tournament.sheetnames)): 
        tur=tournament.worksheets[r]
        opr=wb.worksheets[r]
        num=num_list[r]
        init_set(num)



    wb.create_sheet(index=0, title="順番")
    wb.create_sheet(index=1, title="試合番号表")
    wb.create_sheet(index=2, title="コート番号表")
    rot=wb.worksheets[0]
    game=wb.worksheets[1]
    court=wb.worksheets[2]
    copy_list = copy.deepcopy(num_list)
    for i in range(4,-2,-1):
        for r in range(len(tournament.sheetnames)-1,-1,-1):
            opr=wb.worksheets[r+3]
            rank=wb.sheetnames[r+3][0:2]
            if num_list[r]==i:
                game_set(copy_list[r],num_list[r],rank)
                num_list[r]=i-1
    for i in range(1,game.max_row+1):
        row=rot.max_row+1
        if game['E'+str(i)].value!=None and game['G'+str(i)].value!=None:
                rot['A'+str(row)]=game['A'+str(i)].value
                rot['B'+str(row)]=game['D'+str(i)].value
                rot['C'+str(row)]=game['E'+str(i)].value
                rot['D'+str(row)]=game['F'+str(i)].value
                rot['E'+str(row)]=game['G'+str(i)].value
                rot['F'+str(row)]=game['H'+str(i)].value

    for i in range(1,court_number+1):
        court['A'+str(i)].value=i
        court['B'+str(i)]=rot['A'+str(2)].value
        court['C'+str(i)]=rot['B'+str(2)].value
        court['D'+str(i)]=rot['C'+str(2)].value
        court['E'+str(i)]=rot['D'+str(2)].value
        court['J'+str(i)]=rot['E'+str(2)].value
        court['K'+str(i)]=rot['F'+str(2)].value
        for r in range(1,game.max_row+1): 
            if rot['B'+str(2)].value==game['D'+str(r)].value:
                game['I'+str(r)].value="済"
                book=openpyxl.load_workbook("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(court['C'+str(i)].value)+".xlsx")
                score=book.worksheets[0]
                score["N4"]=court['D'+str(i)].value
                score["B11"]=court['D'+str(i)].value
                score["B19"]=court['D'+str(i)].value
                score["B27"]=court['D'+str(i)].value
                score["AG4"]=court['J'+str(i)].value
                score["B13"]=court['J'+str(i)].value
                score["B21"]=court['J'+str(i)].value
                score["B29"]=court['J'+str(i)].value
                score["A4"]=court['B'+str(i)].value
                score["E4"]=court['C'+str(i)].value
                score["AR4"]=court['A'+str(i)].value
                if str(court['E'+str(i)].value)[-2]=='M' or str(court['E'+str(i)].value)[-2]=='D':
                    score["N8"]=str(court['E'+str(i)].value)[1:-3]
                    score["V8"]=str(court['E'+str(i)].value)[-4:-2]
                else:
                    score["N8"]=str(court['E'+str(i)].value)[1:-2]
                    score["V8"]=str(court['E'+str(i)].value)[-2]
                if str(court['E'+str(i)].value)[-2]=='M' or str(court['E'+str(i)].value)[-2]=='D':
                    score["AG8"]=str(court['K'+str(i)].value)[1:-3]
                    score["AO8"]=str(court['K'+str(i)].value)[-4:-2]
                else:
                    score["AG8"]=str(court['K'+str(i)].value)[1:-2]
                    score["AO8"]=str(court['K'+str(i)].value)[-2]
                book.save("C:\\Users\\mech-user\\Desktop\\IoT\\スコア表\\"+str(court['C'+str(i)].value)+".xlsx")

        rot.delete_rows(2)



    tournament.save("トーナメント.xlsx")
    wb.save('管理表.xlsx')
