# -*- coding: utf-8 -*-
import ftplib
import os
import openpyxl

def main(path):
    if os.path.exists('C:\\Users\\mech-user\\Desktop\\IoT\\photo\\'+path+'.png'):
        host_name = "rally-roman.com" # 接続先サーバーのホスト名
        upload_src_path = "C:\\Users\\mech-user\\Desktop\\IoT\\photo\\"+path+".png" # アップロードするファイルパス
        upload_dst_path = "STOR /photo/"+path+".png" # アップロード先のファイルパス
        username = "ishii@rally-roman.com" # サーバーのユーザー名
        password = "yoshiishii009" # サーバーのログインパスワード

        ftp = ftplib.FTP(host_name)
        ftp.set_pasv("true")
        ftp.login(username, password)
        fp = open(upload_src_path, 'rb')
        ftp.storbinary(upload_dst_path ,fp)

        # 終了処理
        ftp.close()
        fp.close()

"""
tournament=openpyxl.load_workbook("トーナメント.xlsx")
for i in range(len(tournament.sheetnames)):
    main('tournament{}'.format(tournament.sheetnames[i][0:2]))
"""

