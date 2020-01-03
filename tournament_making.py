#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import Workbook
import numpy as np
from openpyxl import utils
import copy


#シートの読み込み
def sheet_loading(self):
    global rank_date,sheet
    rank_date=openpyxl.load_workbook(self)
    sheet=rank_date['Sheet1']

#列データから個人のデータをリスト化
def columuns_list_making(self):
    ans_list=[]
    for ans_list_obj in list(sheet.columns)[self]:
            ans_list.append(ans_list_obj.value)
    ans_list.pop(0)
    return ans_list

#日付データの文字列化
def date_list_to_string(self):
    for i in range(len(self)):
        self[i] = utils.datetime.from_excel(self[i])
        self[i]=self[i].strftime('%m%d')
    return self

#リストを種類ごとに並べてソートする
def list_to_category(self):
    self_copy = copy.deepcopy(self)
    self_category=list(set(self_copy))
    self_category.sort()
    return self_category

#リストの要素の足し算
def list_add(self_a,self_b):
    return [x + y for (x, y) in zip(self_a, self_b)]

#リスト同士のすべての組み合わせを作る
def all_combine(self_a,self_b):
    ans_list=[]
    for i in range(len(self_a)):
        for j in range(len(self_b)):
            ans_list.append(self_b[j]+self_a[i])
    return ans_list

#必要数のシート作成
def create_need_sheet(self):
    for i in range(len(self)):
        rank_date.create_sheet(index=i+1,title=self[i])
    return 0


#メンバーリストの作成
def making_memberlist(rank_date_category,rank_date_list,name,circle,grade,rank,date,place,category):
    for i in range(len(rank_date_category)):
        sheet5=rank_date[rank_date_category[i]]
        sheet5["A1"]="ランク"
        sheet5["B1"]="日付"
        sheet5["C1"]="場所"
        sheet5["D1"]="種目"
        sheet5["E1"]="名前"
        sheet5["F1"]="サークル名"
        sheet5["G1"]="学年"
        sheet5["H1"]="強さ"
        for j in range(len(rank_date_list)):
            a=sheet5.max_row
            if rank_date_list[j]==rank_date_category[i]:
                sheet5.cell(row=a+1,column=5).value = name[j]
                sheet5.cell(row=a+1,column=6).value = circle[j]
                sheet5.cell(row=a+1,column=7).value = grade[j]
                sheet5.cell(row=a+1,column=8).value = 0
                sheet5["A2"]=rank[j]
                sheet5["B2"]=date[j]
        if sheet5.max_row==1:
                rank_date.remove(rank_date[rank_date_category[i]])
        sheet5["D2"]=category
        sheet5["C2"]=place[i]

#memberlistの作成
def make_rank_date_list(self):
    sheet_loading(self)
    rank=columuns_list_making(0)
    date=columuns_list_making(1)
    name=columuns_list_making(2)
    circle=columuns_list_making(3)
    grade=columuns_list_making(4)
    date=date_list_to_string(date)
    date_category=list_to_category(date)
    rank_category=list_to_category(rank)
    rank_date_list=list_add(rank,date)
    return all_combine(date_category,rank_category)


#memberlistの作成
def main_memberlist(self,category,place):
    sheet_loading(self)
    rank=columuns_list_making(0)
    date=columuns_list_making(1)
    name=columuns_list_making(2)
    circle=columuns_list_making(3)
    grade=columuns_list_making(4)
    date=date_list_to_string(date)
    date_category=list_to_category(date)
    rank_category=list_to_category(rank)
    rank_date_list=list_add(rank,date)
    rank_date_category=all_combine(date_category,rank_category)
    create_need_sheet(rank_date_category)
    making_memberlist(rank_date_category,rank_date_list,name,circle,grade,rank,date,place,category)
    rank_date.remove(rank_date['Sheet1'])
    rank_date.save('memberlist.xlsx')


def member_play(file):
    #formatの読み込み#
    global game_number
    book = openpyxl.load_workbook("tournament.xlsx")
    member = openpyxl.load_workbook('memberlist.xlsx')
    bup = openpyxl.load_workbook("bup.xlsx")
    sheet2 = member.worksheets[0]
    predate=sheet2["B2"].value


    for r in range(len(member.sheetnames)):
    #参加者リストの読み込み
    #基本情報入力 ランク　場所　種目

        sheet2 = member.worksheets[r]
        rank=sheet2["A2"].value
        date=sheet2["B2"].value
        place=sheet2["C2"].value
        category=sheet2["D2"].value

        #名前・サークル・学年・強さの読み込み
        name_list=[]
        for name_obj in list(sheet2.columns)[4]:
            name_list.append(name_obj.value)

        name_list.pop(0)
        circle_list=[]
        for circle_obj in list(sheet2.columns)[5]:
            circle_list.append(circle_obj.value)
        circle_list.pop(0)
        grade_list=[]
        for grade_obj in list(sheet2.columns)[6]:
            if grade_obj.value=="M1":
                grade_obj.value=5
            elif grade_obj.value=="M2":
                grade_obj.value=6
            elif grade_obj.value=="D1":
                grade_obj.value=7
            elif grade_obj.value=="D2":
                grade_obj.value=8
            grade_list.append(grade_obj.value)
        grade_list.pop(0)
        strength_list=[]
        for strength_obj in list(sheet2.columns)[7]:
            strength_list.append(strength_obj.value)
        strength_list.pop(0)

        #入力情報のリストを作る　名前　サークル　学年
        write_name_list=[]
        write_belong_list=[]
        new_grade_list5=[]
        new_strength_list5=[]
        new_name_list5=[]
        new_circle_list5=[]

        if category=='シングルス':
            for i in range(len(name_list)):
                write_name_list.append(name_list[i])
                write_belong_list.append("("+circle_list[i]+str(grade_list[i])+")")

        else:
            for i in range(0,len(name_list)-1,2):
                write_name_list.append(name_list[i]+"・"+name_list[i+1])
                write_belong_list.append("("+circle_list[i]+str(grade_list[i])+"・"+circle_list[i+1]+str(grade_list[i+1])+")")
                new_name_list5.append(name_list[i]+name_list[i+1])
                new_circle_list5.append(circle_list[i])
                new_grade_list5.append(grade_list[i]+grade_list[i+1])
                new_strength_list5.append(strength_list[i]+strength_list[i+1])
            name_list.clear()
            circle_list.clear()
            grade_list.clear()
            strength_list.clear()
            name_list=new_name_list5
            circle_list=new_circle_list5
            grade_list=new_grade_list5
            strength_list=new_strength_list5



        #学年と強さからポイントをつけて行く
        point_list = [x*5 + y for (x, y) in zip(grade_list, strength_list)]


        #ポイント順にソートした後のindexを返す
        sortindex = np.argsort(point_list)
        new_name_list=[]
        new_circle_list=[]
        new_grade_list=[]
        new_writename_list=[]
        new_writebelong_list=[]
        for i in range(len(sortindex)):
            new_name_list.append(name_list[sortindex[i]])
            new_circle_list.append(circle_list[sortindex[i]])
            new_grade_list.append(grade_list[sortindex[i]])
            new_writename_list.append(write_name_list[sortindex[i]])
            new_writebelong_list.append(write_belong_list[sortindex[i]])
        new_name_list.reverse()
        new_circle_list.reverse()
        new_grade_list.reverse()
        new_writename_list.reverse()
        new_writebelong_list.reverse()
        for i in range(len(sortindex)):
            name_list[i]=new_name_list[i]
            circle_list[i]=new_circle_list[i]
            grade_list[i]=new_grade_list[i]
            write_name_list[i]=new_writename_list[i]
            write_belong_list[i]=new_writebelong_list[i]


        n=sheet2.max_row-1


        if n<=8:
            ws = book.worksheets[0]
            book.copy_worksheet(ws)
            sheet = book.worksheets[4+r]
            tnm=4

        if n>=9 and n<=16:
            ws = book.worksheets[1]
            book.copy_worksheet(ws)
            sheet = book.worksheets[4+r]
            tnm=8
        if n>=17 and n<=32:
            ws = book.worksheets[2]
            book.copy_worksheet(ws)
            sheet = book.worksheets[4+r]
            tnm=16
        if n>=33 and n<=64:
            ws = book.worksheets[3]
            book.copy_worksheet(ws)
            sheet = book.worksheets[4+r]
            tnm=32

        bind=1
        cnt=0
        while True:
            if bind>=n:
                break
            bind *=2
            cnt+=1
        if cnt==6:
            right_index='N'
            right_index2='M'
        if cnt==5:
            right_index='L'
            right_index2='K'
        if cnt==4:
            right_index='J'
            right_index2='I'
        if cnt==3:
            right_index='H'
            right_index2='G'

        print(cnt)

        seed=2**cnt-n

        seed_list=[]
        number_list=[0]*(n+10)

        if seed!=0:
            up_list=[0]*tnm
            div=int(tnm/seed)
            rem=tnm%seed
            ret=min(seed-rem,rem)
            if ret==rem:
                add_list=[div]*seed
            else:
                add_list=[div+1]*seed

            for i in range(ret):
                if ret==rem:
                    add_list[int(seed/ret)*(i+1)-1]=div+1
                else:
                    add_list[int(seed/ret)*(i+1)-1]=div
            counter=-1
            for i in range(seed):
                counter+=add_list[i]
                up_list[counter]=1
            counter=0
            print(tnm)
            for i in range(tnm):
                if up_list[i]==0:
                    counter+=2
                else:
                    counter+=1
                    seed_list.append(counter)
                    number_list[counter-1]=1
            print(seed_list)

        cntseed=0
        for i in range(1,int((n+1)/2)+1):
            for j in range(seed):
                if seed_list[j]==i:
                    cntseed+=1

        if (int((n+1)/2)-cntseed)%2==0:
            left_number=int((n+1)/2)
            if left_number==cntseed:
                left_number-=1
            right_number=n-left_number

        else:
            left_number=int((n+1)/2)-1
            right_number=n-left_number


        #バップ表の取得
        sheet3 = bup['バップリスト']
        max_column=bup['バップリスト'].max_column

        weight_list=[[0 for i in range(len(circle_list))] for j in range(len(circle_list))]
        block=[]
        counting=0
        for i in range(len(circle_list)):
            if number_list[i]==1:
                counting+=2
            else:
                counting+=1
            if counting==4:
                counting=0
                block.append(i)


        seed_name_list=[]
        seed_circle_list=[]
        seed_grade_list=[]
        seed_writename_list=[]
        seed_writebelong_list=[]
        for i in range(len(seed_list)):
            seed_name_list.append(name_list[i])
            seed_circle_list.append(circle_list[i])
            seed_grade_list.append(grade_list[i])
            seed_writename_list.append(write_name_list[i])
            seed_writebelong_list.append(write_belong_list[i])
            new_name_list.pop(0)
            new_grade_list.pop(0)
            new_circle_list.pop(0)
            new_writename_list.pop(0)
            new_writebelong_list.pop(0)





        def blockcount(x):
            for i in range(len(block)):
                if x<=block[i]:
                    break
            return i+1


        def distance(x,y):
            if int(blockcount(x)/8)!=int(blockcount(y)/8):
                return 1
            elif int(blockcount(x)/4)!=int(blockcount(y)/4):
                return 3
            elif int(blockcount(x)/2)!=int(blockcount(y)/2):
                return 6
            else:
                return 10


        bap_list=[]
        for i in range(max_column):
            for bap_obj in list(sheet3.columns)[i]:
                if bap_obj.value!=None:
                    bap_list.append(bap_obj.value)


        #幾つのバップに入っているかをカウント
        cnt_bap_list=[0]*len(new_circle_list)
        for i in range(len(new_circle_list)):
            for j in range(len(bap_list)):
                if new_circle_list[i]==bap_list[j]:
                    cnt_bap_list[i]+=1


        #ポイント順にソートした後のindexを返す
        sortindex = np.argsort(cnt_bap_list)

        new_name_list2=[]
        new_circle_list2=[]
        new_grade_list2=[]
        new_writename_list2=[]
        new_writebelong_list2=[]
        for i in range(len(sortindex)):
            new_name_list2.append(new_name_list[sortindex[i]])
            new_circle_list2.append(new_circle_list[sortindex[i]])
            new_grade_list2.append(grade_list[sortindex[i]])
            new_writename_list2.append(new_writename_list[sortindex[i]])
            new_writebelong_list2.append(new_writebelong_list[sortindex[i]])
        for i in range(len(sortindex)):
            new_name_list[i]=new_name_list2[i]
            new_circle_list[i]=new_circle_list2[i]
            new_grade_list[i]=new_grade_list2[i]
            new_writename_list[i]=new_writename_list2[i]
            new_writebelong_list[i]=new_writebelong_list2[i]
        new_name_list.reverse()
        new_circle_list.reverse()
        new_grade_list.reverse()
        new_writename_list.reverse()
        new_writebelong_list.reverse()




        new_new_name_list=seed_name_list+new_name_list
        new_new_circle_list=seed_circle_list+new_circle_list
        new_new_grade_list=seed_grade_list+new_grade_list
        new_new_circle_list=seed_circle_list+new_circle_list
        new_new_writename_list=seed_writename_list+new_writename_list
        new_new_writebelong_list=seed_writebelong_list+new_writebelong_list



        number_decide=[]
        for i in range(len(seed_list)):
            ans=1000000
            for j in range(len(new_new_circle_list)):
                for k in range(len(seed_list)):
                    if j==seed_list[k]-1:
                        if ans>weight_list[i][j]:
                            ans=weight_list[i][j]
                            y=j
            number_decide.append(y)
            for j in range(len(new_new_circle_list)):
                weight_list[j][y]+=10000
            list_bap=[0]*max_column
            weight_circle=[]
            for j in range(max_column):
                for bap_obj in list(sheet3.columns)[j]:
                    if bap_obj.value!=None:
                        if new_new_circle_list[i]==bap_obj.value:
                            list_bap[j]=1
            for j in range(max_column):
                if list_bap[j]==1:
                    for bap_obj in list(sheet3.columns)[j]:
                        if bap_obj.value!=None:
                            weight_circle.append(bap_obj.value)
            for j in range(len(new_new_circle_list)):
                for k in range(len(weight_circle)):
                    if weight_circle[k]==new_new_circle_list[j] and i!=j:
                        for l in range(len(circle_list)):
                            weight_list[j][l]+=distance(y,l)
            for j in range(len(new_new_circle_list)):
                if new_new_circle_list[i]==new_new_circle_list[j] and j!=i:
                    for k in range(len(circle_list)):
                        weight_list[j][k]+=5*distance(y,k)


        for i in range(len(seed_list),len(new_new_circle_list)):
            ans=1000000
            for j in range(len(new_new_circle_list)):
                if ans>weight_list[i][j]:
                    ans=weight_list[i][j]
                    y=j
            number_decide.append(y)
            for j in range(len(new_new_circle_list)):
                weight_list[j][y]+=10000
            list_bap=[0]*max_column
            weight_circle=[]
            for j in range(max_column):
                for bap_obj in list(sheet3.columns)[j]:
                    if bap_obj.value!=None:
                        if new_new_circle_list[i]==bap_obj.value:
                            list_bap[j]=1
            for j in range(max_column):
                if list_bap[j]==1:
                    for bap_obj in list(sheet3.columns)[j]:
                        if bap_obj.value!=None:
                            weight_circle.append(bap_obj.value)
            for j in range(len(new_new_circle_list)):
                for k in range(len(weight_circle)):
                    if weight_circle[k]==new_new_circle_list[j]:
                        for l in range(len(circle_list)):
                            weight_list[j][l]+=distance(y,l)
            for j in range(len(new_new_circle_list)):
                if new_new_circle_list[i]==new_new_circle_list[j]:
                    for k in range(len(circle_list)):
                        weight_list[j][k]+=5*distance(y,k)



        new_name_list3=[0]*len(number_decide)
        new_circle_list3=[0]*len(number_decide)
        new_grade_list3=[0]*len(number_decide)
        new_writename_list3=[0]*len(number_decide)
        new_writebelong_list3=[0]*len(number_decide)

        for i in range(len(number_decide)):
            new_name_list3[number_decide[i]]=new_new_name_list[i]
            new_circle_list3[number_decide[i]]=new_new_circle_list[i]
            new_grade_list3[number_decide[i]]=new_new_grade_list[i]
            new_writename_list3[number_decide[i]]=new_new_writename_list[i]
            new_writebelong_list3[number_decide[i]]=new_new_writebelong_list[i]

        for i in range(len(number_decide)):
            name_list[i]=new_name_list3[i]
            circle_list[i]=new_circle_list3[i]
            grade_list[i]=new_grade_list3[i]
            write_name_list[i]=new_writename_list3[i]
            write_belong_list[i]=new_writebelong_list3[i]

        print(date)
        print(predate)
        if date!=predate or r==0:
            print("あ")
            game_number=100
        else:
            game_number=100+int(game_number/100)*100
        predate=date
        # 罫線(外枠)を設定
        right = Border(right=Side(style='thin', color='000000'))
        left = Border(left=Side(style='thin', color='000000'))
        top_right = Border(top=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000')
        )
        under_right = Border(bottom=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000')
        )
        top_left = Border(top=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000')
        )
        under_left = Border(bottom=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000')
        )
        under=Border(bottom=Side(style='thin', color='000000'))

        count_left=8
        count_right=8
        noseed=-1

        if cnt>=3:
            for i in range(n):
                if left_number>=i+1:
                    if number_list[i]==1:
                        sheet['A'+str(count_left)].value=write_name_list[i]
                        sheet['A'+str(count_left+1)].value=write_belong_list[i]
                        sheet['B'+str(count_left)].border=under
                        count_left+=8
                    elif noseed==-1 and number_list[i]==0:
                        sheet['A'+str(count_left-2)].value=write_name_list[i]
                        sheet['A'+str(count_left-1)].value=write_belong_list[i]
                        sheet['B'+str(count_left-1)].border=top_right
                        sheet['B'+str(count_left)].border=right
                        sheet.merge_cells('B'+str(count_left)+':'+'B'+str(count_left+1))
                        sheet['B'+str(count_left)].value=game_number
                        sheet['B'+str(count_left)].border=right
                        game_number+=1
                        noseed*=-1
                    else:
                        sheet['A'+str(count_left+2)].value=write_name_list[i]
                        sheet['A'+str(count_left+3)].value=write_belong_list[i]
                        sheet['B'+str(count_left+2)].border=under_right
                        sheet['B'+str(count_left+1)].border=right
                        count_left+=8
                        noseed*=-1
                else:
                    if number_list[i]==1:
                        sheet[right_index+str(count_right)].value=write_name_list[i]
                        sheet[right_index+str(count_right+1)].value=write_belong_list[i]
                        sheet[right_index2+str(count_right)].border=under
                        count_right+=8
                    elif noseed==-1 and number_list[i]==0:
                        sheet[right_index+str(count_right-2)].value=write_name_list[i]
                        sheet[right_index+str(count_right-1)].value=write_belong_list[i]
                        sheet[right_index2+str(count_right-1)].border=top_left
                        sheet[right_index2+str(count_right)].border=left
                        sheet.merge_cells(right_index2+str(count_right)+':'+right_index2+str(count_right+1))
                        sheet[right_index2+str(count_right)].value=game_number
                        sheet[right_index2+str(count_right)].border=left
                        game_number+=1
                        noseed*=-1
                    else:
                        sheet[right_index+str(count_right+2)].value=write_name_list[i]
                        sheet[right_index+str(count_right+3)].value=write_belong_list[i]
                        sheet[right_index2+str(count_right+2)].border=under_left
                        sheet[right_index2+str(count_right+1)].border=left
                        count_right+=8
                        noseed*=-1



        if cnt==3:
            sheet["C12"]=game_number
            game_number+=1
            sheet["F12"]=game_number
            game_number+=1
            sheet["D13"]=game_number
            game_number+=1
            sheet["C1"]=rank
            sheet["C3"]=date
            sheet["C4"]=place

        elif cnt==4:
            sheet["C12"]=game_number
            game_number+=1
            sheet["C28"]=game_number
            game_number+=1
            sheet["H12"]=game_number
            game_number+=1
            sheet["H28"]=game_number
            game_number+=1
            sheet["D20"]=game_number
            game_number+=1
            sheet["G20"]=game_number
            game_number+=1
            sheet["E21"]=game_number
            game_number+=1
            sheet["D1"]=rank
            sheet["D3"]=date
            sheet["D4"]=place

        elif cnt==5:
            sheet["C12"]=game_number
            game_number+=1
            sheet["C28"]=game_number
            game_number+=1
            sheet["C44"]=game_number
            game_number+=1
            sheet["C60"]=game_number
            game_number+=1
            sheet["J12"]=game_number
            game_number+=1
            sheet["J28"]=game_number
            game_number+=1
            sheet["J44"]=game_number
            game_number+=1
            sheet["J60"]=game_number
            game_number+=1
            sheet["D20"]=game_number
            game_number+=1
            sheet["D52"]=game_number
            game_number+=1
            sheet["I20"]=game_number
            game_number+=1
            sheet["I52"]=game_number
            game_number+=1
            sheet["E36"]=game_number
            game_number+=1
            sheet["H36"]=game_number
            game_number+=1
            sheet["F37"]=game_number
            game_number+=1
            sheet["E1"]=rank
            sheet["E3"]=date
            sheet["E4"]=place

        else:
            sheet["C12"]=game_number
            game_number+=1
            sheet["C28"]=game_number
            game_number+=1
            sheet["C44"]=game_number
            game_number+=1
            sheet["C60"]=game_number
            game_number+=1
            sheet["C76"]=game_number
            game_number+=1
            sheet["C92"]=game_number
            game_number+=1
            sheet["C108"]=game_number
            game_number+=1
            sheet["C124"]=game_number
            game_number+=1
            sheet["L12"]=game_number
            game_number+=1
            sheet["L28"]=game_number
            game_number+=1
            sheet["L44"]=game_number
            game_number+=1
            sheet["L60"]=game_number
            game_number+=1
            sheet["L76"]=game_number
            game_number+=1
            sheet["L92"]=game_number
            game_number+=1
            sheet["L108"]=game_number
            game_number+=1
            sheet["L124"]=game_number
            game_number+=1
            sheet["D20"]=game_number
            game_number+=1
            sheet["D52"]=game_number
            game_number+=1
            sheet["D84"]=game_number
            game_number+=1
            sheet["D116"]=game_number
            game_number+=1
            sheet["K20"]=game_number
            game_number+=1
            sheet["K52"]=game_number
            game_number+=1
            sheet["K84"]=game_number
            game_number+=1
            sheet["K116"]=game_number
            game_number+=1
            sheet["E36"]=game_number
            game_number+=1
            sheet["E100"]=game_number
            game_number+=1
            sheet["J36"]=game_number
            game_number+=1
            sheet["J100"]=game_number
            game_number+=1
            sheet["F68"]=game_number
            game_number+=1
            sheet["I68"]=game_number
            game_number+=1
            sheet["G69"]=game_number
            game_number+=1
            sheet["F1"]=rank
            sheet["F3"]=date
            sheet["F4"]=place

        for i in range(1000):
            sheet['A'+str(i+1)].alignment = Alignment(
                horizontal='center',
                vertical='bottom',
            )
            sheet[right_index+str(i+1)].alignment = Alignment(
                horizontal='center',
                vertical='bottom',
            )

        sheet.title=rank+date
    book.remove(book['5-8人用'])
    book.remove(book['9-16人用'])
    book.remove(book['17-32人用'])
    book.remove(book['33-64人用'])
    book.save(file)
def main(self,category,place,file):
    main_memberlist(self,category,place)
    member_play(file)
if __name__ == '__main__':
    main_memberlist(file)
    member_play()
